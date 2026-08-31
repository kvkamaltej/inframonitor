from collections import deque
from collections.abc import Iterator
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from typing import Any
from urllib.parse import quote

import io
import json
import os
import posixpath
import re
import secrets
import tempfile
import threading
import uuid

import asyncio

from jose import jwt

from sqlalchemy import delete, func, select, update
from sqlalchemy.exc import IntegrityError
from fastapi import APIRouter, Body, Depends, File, Form, Header, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect, status
from fastapi.responses import FileResponse, StreamingResponse
from starlette.background import BackgroundTask
from starlette.responses import RedirectResponse
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.database import SessionLocal, get_db
from app.core.crypto import decrypt_secret, encrypt_secret
from app.core.menus import DEFAULT_ROLE_MENUS, MENU_ITEMS, ROLE_KEYS, ROLE_MENUS_KEY, menus_for_role
from app.core.security import SEEDED_GUEST_EMAIL, create_access_token, decode_token, guest_claims, hash_password, is_loopback_client, require_admin, require_admin_not_guest, require_admin_or_developer, require_user, verify_password
from app.services.secrets import VaultError, get_vault_config, load_credentials, save_vault_config, store_credentials, test_vault
from app.services import alert_rules
from app.services import db_backend
from app.services import db_backup
from app.services import oidc
from app.models.entities import AccessPolicy, AppSetting, AuditLog, DbConnection, DbQueryHistory, Folder, KubeCluster, Server, ServerStatus, ShellFavorite, User, UserPolicyAssignment, UserServerAccess
from app.schemas.contracts import (
    AccessPolicyCreate,
    AccessPolicyRead,
    AppDbConfigRead,
    AppDbConnectionRequest,
    AppDbMigrateResult,
    AppDbTestResult,
    AlertBufferResponse,
    AlertRecord,
    AlertWebhookResult,
    ConnectionResult,
    ContainerRead,
    ImageRead,
    CredentialPayload,
    DbConnectionCreate,
    DbConnectionQueryRequest,
    DbConnectionRead,
    DbConnectionRequest,
    DbConnectionResult,
    DbConnectionUpdate,
    DbColumn,
    DbConstraint,
    DbDatabase,
    DbForeignKey,
    DbGenerateSqlRequest,
    DbGenerateSqlResult,
    DbIndex,
    DbQueryHistoryRead,
    DbQueryRequest,
    DbQueryResult,
    DbRoutine,
    DbSchema,
    DbSchemaRenameRequest,
    DbTable,
    FolderCreate,
    FolderRead,
    IntegrationStatus,
    ActionResult,
    AlertRuleCreate,
    KubeClusterCreate,
    KubeClusterRead,
    KubeClusterUpdate,
    KubeLogShippingUpdate,
    KubeCordonRequest,
    KubeDeployment,
    KubeEvent,
    KubeHealth,
    KubeNode,
    KubeOverview,
    KubePod,
    KubePodLogs,
    KubeScaleRequest,
    KubeTestResult,
    LoginRequest,
    LogResponse,
    LogShippingUpdate,
    LogSourceEntry,
    MeResponse,
    MonitoringInstallRequest,
    OperationRequest,
    OptionCreate,
    OptionList,
    PasswordChange,
    ProfileUpdate,
    PolicyAssignmentUpdate,
    PrivilegedOperationResult,
    RoleMenusResponse,
    RoleMenusUpdate,
    ServerCreate,
    ServerFolderUpdate,
    ServerImportRequest,
    ServerImportResult,
    ServerMonitoringState,
    ServerRead,
    ServerUpdate,
    ServerAccessUpdate,
    ServiceLogRequest,
    ServiceRestartRequest,
    SftpDeleteRequest,
    SftpDeleteResult,
    SftpListing,
    SftpUploadResult,
    ShellFavoriteCreate,
    ShellFavoriteRead,
    Summary,
    TokenResponse,
    TomcatActionRequest,
    TomcatInstance,
    TomcatLogRequest,
    UserCreate,
    UserRead,
    UserUpdate,
    UserServerAccessRead,
    VaultConfigRead,
    VaultConfigWrite,
    VaultTestResult,
    WarDeployResult,
)
from app.services.csv_import import CsvImportError, import_servers
from app.services.db_console import DEFAULT_PORTS, ENGINES, DbConsoleError
from app.services import db_console
from app.services import db_metadata
from app.services import kube
from app.services.kube import KubeError
from app.services.integrations import check_integrations
from app.services import monitoring
from app.services.monitoring import MonitoringError
from app.services import node_exporter
from app.services.node_exporter import NodeExporterError
from app.services.inventory import InventoryService, to_read
from app.services.validation import validate_host_address
from app.services.ssh_ops import (
    ShellSession,
    SftpTooLarge,
    SshOperationError,
    SudoPasswordRequired,
    container_env_file,
    container_logs,
    deploy_war,
    detect_log_capabilities,
    discover_host,
    discover_tomcat,
    list_containers_with_ports,
    list_images,
    probe_vitals,
    remove_container,
    remove_image,
    restart_container,
    restart_service,
    run_command,
    service_logs,
    sftp_delete,
    sftp_download,
    sftp_list,
    sftp_stat,
    sftp_upload,
    tomcat_action,
    tomcat_logs,
)

router = APIRouter()


def _role_value(role: str) -> str:
    normalized = role.strip().lower()
    if normalized == "administrator":
        normalized = "admin"
    if normalized not in {"admin", "developer", "support", "guest"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Role must be admin, developer, support, or guest")
    return normalized


def _setting_list(db: Session, key: str) -> list[str]:
    setting = db.get(AppSetting, key)
    if not setting:
        return []
    try:
        parsed = json.loads(setting.value)
        return parsed if isinstance(parsed, list) else []
    except json.JSONDecodeError:
        return []


def _save_setting_list(db: Session, key: str, values: list[str]) -> list[str]:
    clean = []
    for value in values:
        item = value.strip().lower()
        if item and item not in clean:
            clean.append(item)
    setting = db.get(AppSetting, key)
    if not setting:
        setting = AppSetting(key=key, value="[]")
        db.add(setting)
    setting.value = json.dumps(clean)
    db.commit()
    return clean


def _csv(values: list[str]) -> str:
    return ",".join(sorted({value.strip().lower() for value in values if value.strip()}))


def _csv_split(value: str) -> list[str]:
    return [item for item in value.split(",") if item]


def _current_user(db: Session, claims: dict) -> User:
    user = db.scalar(select(User).where(User.email == claims.get("sub")))
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return user


def _sees_all_servers(claims: dict) -> bool:
    # Admins AND developers get full inventory visibility + access; support (and any other
    # non-privileged role) sees only servers granted via UserServerAccess / access policies.
    # A developer provisioned via Keycloak has no per-server grants, so gating them the same as
    # support left them with ZERO servers -- developers are trusted with the whole inventory, and
    # role-level checks (require_admin / require_admin_or_developer) still gate dangerous actions.
    return claims.get("role") in {"admin", "administrator", "developer"}


def _policy_applies(policy: AccessPolicy, server: Server) -> bool:
    environments = set(_csv_split(policy.environments))
    server_types = set(_csv_split(policy.server_types))
    tags = set(_csv_split(policy.tags))
    server_ids = set(_csv_split(policy.server_ids))
    server_tags = {tag.strip().lower() for tag in server.tags.split(",") if tag.strip()}
    return (
        server.public_id.lower() in server_ids
        or server.environment.lower() in environments
        or server.server_type.lower() in server_types
        or bool(tags.intersection(server_tags))
    )


def _accessible_server_ids(db: Session, user: User) -> set[int]:
    direct_ids = set(db.scalars(select(UserServerAccess.server_id).where(UserServerAccess.user_id == user.id)).all())
    policies = db.scalars(
        select(AccessPolicy)
        .join(UserPolicyAssignment, UserPolicyAssignment.policy_id == AccessPolicy.id)
        .where(UserPolicyAssignment.user_id == user.id)
    ).all()
    if not policies:
        return direct_ids
    servers = db.scalars(select(Server)).all()
    for server in servers:
        if any(_policy_applies(policy, server) for policy in policies):
            direct_ids.add(server.id)
    return direct_ids


def _server_or_404(db: Session, server_id: str, claims: dict | None = None) -> Server:
    server = db.scalar(select(Server).where(Server.public_id == server_id))
    if not server and server_id.isdigit():
        server = db.get(Server, int(server_id))
    if not server:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Server not found")
    if claims and not _sees_all_servers(claims):
        user = _current_user(db, claims)
        if server.id not in _accessible_server_ids(db, user):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Server is not assigned to this user")
    return server


def _local_login_allowed() -> bool:
    # Local username/password login is available UNLESS Keycloak is configured -- then web sign-in
    # is Keycloak-only, and local login is off unless the ALLOW_LOCAL_LOGIN break-glass is set. With
    # no OIDC configured (standalone/SQLite/desktop) local login is always the way in.
    settings = get_settings()
    return (not oidc.configured()) or settings.allow_local_login


@router.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> TokenResponse:
    if not _local_login_allowed():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Local login is disabled. Sign in with Keycloak.",
        )
    user = db.scalar(select(User).where(User.email == payload.email))
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    return TokenResponse(access_token=create_access_token(user.email, user.role.value))


# --- optional Keycloak / OIDC single sign-on (Authorization Code + PKCE, BFF) -------------------
# All three routes are PUBLIC (no auth dependency) and inert when OIDC is unconfigured: /status
# reports enabled:false and /login redirects back to the SPA with #oidc_error=disabled. The state,
# PKCE verifier, and nonce are carried across the round-trip in a short-lived signed cookie so the
# backend stays stateless. Everything below is additive -- /auth/login above is unchanged.
_OIDC_FLOW_COOKIE = "oidc_flow"
_OIDC_ID_COOKIE = "oidc_id"  # id_token stashed for RP-initiated logout (id_token_hint)
_OIDC_COOKIE_PATH = "/api/auth/oidc"
_OIDC_FLOW_MAX_AGE = 300  # 5 minutes


def _oidc_app_base() -> str:
    return get_settings().app_public_url.rstrip("/")


def _oidc_error_redirect(message: str) -> RedirectResponse:
    return RedirectResponse(f"{_oidc_app_base()}/#oidc_error={quote(message)}", status_code=302)


@router.get("/auth/oidc/status")
def oidc_status() -> dict:
    # `local_login` tells the login screen whether to render the email/password form at all: it is
    # hidden when Keycloak is the only permitted sign-in (OIDC configured + break-glass off).
    return {"enabled": oidc.configured(), "local_login": _local_login_allowed()}


@router.get("/auth/oidc/login")
def oidc_login() -> RedirectResponse:
    if not oidc.configured():
        return RedirectResponse(f"{_oidc_app_base()}/#oidc_error=disabled", status_code=302)
    settings = get_settings()
    state = secrets.token_urlsafe(24)
    verifier, challenge = oidc.make_pkce()
    nonce = secrets.token_urlsafe(16)
    flow = jwt.encode(
        {
            "state": state,
            "verifier": verifier,
            "nonce": nonce,
            "exp": datetime.now(timezone.utc) + timedelta(seconds=_OIDC_FLOW_MAX_AGE),
        },
        settings.jwt_secret,
        algorithm="HS256",
    )
    resp = RedirectResponse(oidc.build_authorize_url(state, challenge, nonce), status_code=302)
    resp.set_cookie(
        _OIDC_FLOW_COOKIE,
        flow,
        max_age=_OIDC_FLOW_MAX_AGE,
        httponly=True,
        samesite="lax",
        secure=False,
        path=_OIDC_COOKIE_PATH,
    )
    return resp


@router.get("/auth/oidc/callback")
def oidc_callback(request: Request, db: Session = Depends(get_db)) -> RedirectResponse:
    if not oidc.configured():
        return RedirectResponse(f"{_oidc_app_base()}/#oidc_error=disabled", status_code=302)

    # 1. Recover and verify the signed flow cookie (state + PKCE verifier + nonce).
    cookie = request.cookies.get(_OIDC_FLOW_COOKIE)
    if not cookie:
        return _oidc_error_redirect("expired")
    try:
        flow = jwt.decode(cookie, get_settings().jwt_secret, algorithms=["HS256"])
    except Exception:
        return _oidc_error_redirect("expired")

    # 2. Guard the callback: an IdP-reported error, or a state that does not match, is a hard stop.
    params = request.query_params
    if params.get("error") or params.get("state") != flow.get("state"):
        return _oidc_error_redirect("state")
    code = params.get("code")
    if not code:
        return _oidc_error_redirect("state")

    # 3. Exchange the code, validate the access token, map claims to an app identity.
    try:
        tokens = oidc.exchange_code(code, flow.get("verifier", ""))
        claims = oidc.validate_token(tokens["access_token"], flow.get("nonce"))
        email, role = oidc.extract_identity(claims)
    except oidc.OidcError as exc:
        return _oidc_error_redirect(str(exc))
    except Exception as exc:  # any unexpected shape from the IdP still surfaces cleanly to the SPA
        return _oidc_error_redirect(str(exc))

    # 4. Provision or sync the local user. Keycloak is the source of truth for the role.
    from app.models.entities import Role

    model_role = Role.administrator if role == "admin" else Role(role)
    user = db.scalar(select(User).where(func.lower(User.email) == email.lower()))
    if user is None:
        full_name = str(claims.get("name") or claims.get("preferred_username") or email)
        # Unusable local password: SSO users cannot fall back to /auth/login.
        user = User(
            email=email,
            full_name=full_name,
            password_hash=hash_password(secrets.token_urlsafe(32)),
            role=model_role,
        )
        db.add(user)
    else:
        user.role = model_role
    db.commit()
    db.refresh(user)

    # 5. Mint the app JWT and hand it to the SPA in the redirect fragment; clear the flow cookie.
    app_jwt = create_access_token(user.email, user.role.value)
    resp = RedirectResponse(f"{_oidc_app_base()}/#access_token={app_jwt}", status_code=302)
    resp.delete_cookie(_OIDC_FLOW_COOKIE, path=_OIDC_COOKIE_PATH)
    # Stash the id_token (httponly, same short OIDC path) so /auth/oidc/logout can pass it as
    # id_token_hint and end the Keycloak session cleanly (no logout-confirmation prompt).
    id_token = tokens.get("id_token")
    if id_token:
        resp.set_cookie(
            _OIDC_ID_COOKIE, id_token, max_age=get_settings().access_token_expire_minutes * 60,
            httponly=True, samesite="lax", secure=False, path=_OIDC_COOKIE_PATH,
        )
    return resp


@router.get("/auth/oidc/logout")
def oidc_logout(request: Request) -> RedirectResponse:
    """End the Keycloak SSO session too, then return to the app. Without this, clearing the app's
    own token leaves the IdP session alive and the next sign-in is silent (logout appears broken)."""
    base = _oidc_app_base() or "/"
    if not oidc.configured():
        return RedirectResponse(base, status_code=302)
    id_token = request.cookies.get(_OIDC_ID_COOKIE)
    try:
        target = oidc.end_session_url(base, id_token)
    except Exception:
        target = base
    resp = RedirectResponse(target, status_code=302)
    resp.delete_cookie(_OIDC_ID_COOKIE, path=_OIDC_COOKIE_PATH)
    return resp


def _configured_default_password() -> str:
    # the seeded default from ADMIN_PASSWORD; empty means "no default to warn about"
    return str(getattr(get_settings(), "admin_password", "") or "")


# bcrypt verification is deliberately slow, and /auth/me is called on every page load, so
# memoise on the stored hash: changing the password changes the hash and misses the cache
@lru_cache(maxsize=64)
def _hash_matches_default(password_hash: str, default_password: str) -> bool:
    if not password_hash or not default_password:
        return False
    try:
        return verify_password(default_password, password_hash)
    except Exception:
        return False


@router.get("/auth/me", response_model=MeResponse)
def me(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> MeResponse:
    matrix = _role_menus_saved(db)
    # Desktop guest: no backing user row to load, so answer directly. Full-access (admin) but
    # flagged guest so the UI can show the Sign-in option; its menus come from the "guest" row.
    if claims.get("guest"):
        return MeResponse(
            email="Guest",
            full_name="Guest",
            role="admin",
            using_default_password=False,
            guest=True,
            menus=menus_for_role(matrix, "admin", guest=True),
        )
    role = claims.get("role")
    # this used to echo the token only; it now loads the user because the default-password
    # warning has to compare against the stored hash
    user = _current_user(db, claims)
    return MeResponse(
        email=claims.get("sub") or user.email,
        full_name=user.full_name,
        role="admin" if role == "administrator" else str(role or ""),
        using_default_password=_hash_matches_default(user.password_hash, _configured_default_password()),
        menus=menus_for_role(matrix, role, guest=False),
    )


@router.patch("/auth/profile", response_model=MeResponse)
def update_profile(payload: ProfileUpdate, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> MeResponse:
    # A desktop guest has no account row to edit.
    if claims.get("guest"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Guest sessions have no profile. Sign in to edit your details.")
    user = _current_user(db, claims)
    user.full_name = payload.full_name.strip()
    db.commit()
    db.refresh(user)
    return me(claims, db)


@router.post("/auth/change-password", response_model=ConnectionResult)
def change_password(payload: PasswordChange, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ConnectionResult:
    user = _current_user(db, claims)
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Current password is incorrect")
    user.password_hash = hash_password(payload.new_password)
    db.commit()
    return ConnectionResult(ok=True, message="Password changed")


@router.get("/users", response_model=list[UserRead])
def list_users(_: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> list[UserRead]:
    users = db.scalars(select(User).order_by(User.email)).all()
    return [
        UserRead(
            id=user.id,
            email=user.email,
            full_name=user.full_name,
            role="admin" if user.role.value == "administrator" else user.role.value,
            created_at=user.created_at,
        )
        for user in users
    ]


@router.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(payload: UserCreate, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> UserRead:
    from app.models.entities import Role

    if db.scalar(select(User).where(User.email == payload.email)):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="User already exists")
    role = _role_value(payload.role)
    model_role = Role.administrator if role == "admin" else Role(role)
    user = User(email=payload.email, full_name=payload.full_name, password_hash=hash_password(payload.password), role=model_role)
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserRead(id=user.id, email=user.email, full_name=user.full_name, role=role, created_at=user.created_at)


@router.patch("/users/{user_id}", response_model=UserRead)
def update_user(user_id: int, payload: UserUpdate, claims: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> UserRead:
    from app.models.entities import Role

    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    data = payload.model_dump(exclude_unset=True)

    if "full_name" in data and data["full_name"] is not None:
        user.full_name = data["full_name"]
    if "role" in data and data["role"] is not None:
        role = _role_value(data["role"])
        # a caller must not change their own role: an admin who demotes themselves would lock the
        # account out of the very screen they are on, so this is refused rather than silently applied.
        current = "admin" if user.role.value == "administrator" else user.role.value
        if user.email == claims.get("sub") and role != current:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="You cannot change your own role")
        user.role = Role.administrator if role == "admin" else Role(role)
    if data.get("password"):
        if len(data["password"]) < 8:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Password must be at least 8 characters")
        user.password_hash = hash_password(data["password"])

    db.commit()
    db.refresh(user)
    return UserRead(
        id=user.id,
        email=user.email,
        full_name=user.full_name,
        role="admin" if user.role.value == "administrator" else user.role.value,
        created_at=user.created_at,
    )


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(user_id: int, claims: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> None:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.email == claims.get("sub"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="You cannot delete your own user")
    # The built-in guest account is a fixture the "guest" menu row and desktop guest state lean
    # on. Blocking the delete keeps it stable rather than having it silently reappear on the next
    # restart (the seeder would recreate it anyway).
    if user.email == SEEDED_GUEST_EMAIL:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The built-in guest user cannot be deleted.")
    db.delete(user)
    db.commit()


@router.get("/users/{user_id}/server-access", response_model=UserServerAccessRead)
def get_user_server_access(user_id: int, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> UserServerAccessRead:
    if not db.get(User, user_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    rows = db.scalars(select(Server.public_id).join(UserServerAccess, UserServerAccess.server_id == Server.id).where(UserServerAccess.user_id == user_id)).all()
    return UserServerAccessRead(user_id=user_id, server_ids=list(rows))


@router.put("/users/{user_id}/server-access", response_model=UserServerAccessRead)
def update_user_server_access(user_id: int, payload: ServerAccessUpdate, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> UserServerAccessRead:
    if not db.get(User, user_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    servers = db.scalars(select(Server).where(Server.public_id.in_(payload.server_ids))).all() if payload.server_ids else []
    found = {server.public_id for server in servers}
    missing = [server_id for server_id in payload.server_ids if server_id not in found]
    if missing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Unknown server ids: {', '.join(missing)}")
    db.execute(delete(UserServerAccess).where(UserServerAccess.user_id == user_id))
    for server in servers:
        db.add(UserServerAccess(user_id=user_id, server_id=server.id))
    db.commit()
    return get_user_server_access(user_id, _, db)


def _policy_read(policy: AccessPolicy) -> AccessPolicyRead:
    return AccessPolicyRead(
        id=policy.id,
        name=policy.name,
        description=policy.description,
        environments=_csv_split(policy.environments),
        server_types=_csv_split(policy.server_types),
        tags=_csv_split(policy.tags),
        server_ids=_csv_split(policy.server_ids),
        assigned_user_ids=[assignment.user_id for assignment in policy.assignments],
    )


@router.get("/policies", response_model=list[AccessPolicyRead])
def list_policies(_: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> list[AccessPolicyRead]:
    return [_policy_read(policy) for policy in db.scalars(select(AccessPolicy).order_by(AccessPolicy.name)).all()]


@router.post("/policies", response_model=AccessPolicyRead, status_code=status.HTTP_201_CREATED)
def create_policy(payload: AccessPolicyCreate, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> AccessPolicyRead:
    if db.scalar(select(AccessPolicy).where(AccessPolicy.name == payload.name.strip())):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Policy already exists")
    policy = AccessPolicy(
        name=payload.name.strip(),
        description=payload.description,
        environments=_csv(payload.environments),
        server_types=_csv(payload.server_types),
        tags=_csv(payload.tags),
        server_ids=_csv(payload.server_ids),
    )
    db.add(policy)
    db.commit()
    db.refresh(policy)
    return _policy_read(policy)


@router.put("/policies/{policy_id}/users", response_model=AccessPolicyRead)
def assign_policy_users(policy_id: int, payload: PolicyAssignmentUpdate, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> AccessPolicyRead:
    policy = db.get(AccessPolicy, policy_id)
    if not policy:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Policy not found")
    users = db.scalars(select(User).where(User.id.in_(payload.user_ids))).all() if payload.user_ids else []
    found = {user.id for user in users}
    missing = [str(user_id) for user_id in payload.user_ids if user_id not in found]
    if missing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Unknown user ids: {', '.join(missing)}")
    db.execute(delete(UserPolicyAssignment).where(UserPolicyAssignment.policy_id == policy_id))
    for user in users:
        db.add(UserPolicyAssignment(user_id=user.id, policy_id=policy.id))
    db.commit()
    db.refresh(policy)
    return _policy_read(policy)


@router.get("/settings/options", response_model=OptionList)
def get_options(_: dict = Depends(require_user), db: Session = Depends(get_db)) -> OptionList:
    return OptionList(
        environments=_setting_list(db, "environments"),
        server_types=_setting_list(db, "server_types"),
        application_types=_setting_list(db, "application_types"),
    )


# --- role -> menu matrix -------------------------------------------------------------------
#
# A per-role map of which sidebar items a role SEES. Persisted as one JSON AppSetting
# ("role_menus"). This governs menu VISIBILITY only -- endpoint authorization is unchanged and
# still enforced by the require_admin / require_admin_not_guest guards, so a role seeing a menu
# item does not gain access to the API behind it.


def _role_menus_saved(db: Session) -> dict[str, list[str]]:
    # Returns the matrix as saved, starting from the defaults so any role absent from the stored
    # JSON is filled in. A role present but set to an empty list is preserved as empty here (the
    # editor must reflect what was saved); the never-blank fallback happens in menus_for_role,
    # which is what /auth/me uses to build a caller's effective list.
    result = {role: list(items) for role, items in DEFAULT_ROLE_MENUS.items()}
    setting = db.get(AppSetting, ROLE_MENUS_KEY)
    if setting:
        try:
            parsed = json.loads(setting.value)
        except json.JSONDecodeError:
            parsed = {}
        if isinstance(parsed, dict):
            for role, items in parsed.items():
                if role in ROLE_KEYS and isinstance(items, list):
                    cleaned: list[str] = []
                    for item in items:
                        if item in MENU_ITEMS and item not in cleaned:
                            cleaned.append(item)
                    result[role] = cleaned
    return result


@router.get("/settings/role-menus", response_model=RoleMenusResponse)
def get_role_menus(_: dict = Depends(require_user), db: Session = Depends(get_db)) -> RoleMenusResponse:
    # require_user: every signed-in caller (and the desktop guest) can READ the matrix so the
    # frontend can render its own menu. Only the editor UI (admins) fetches the whole grid.
    return RoleMenusResponse(roles=list(ROLE_KEYS), items=list(MENU_ITEMS), menus=_role_menus_saved(db))


@router.put("/settings/role-menus", response_model=RoleMenusResponse)
def update_role_menus(payload: RoleMenusUpdate, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> RoleMenusResponse:
    clean: dict[str, list[str]] = {}
    for role, items in payload.menus.items():
        if role not in ROLE_KEYS:
            continue
        deduped: list[str] = []
        for item in items:
            if item in MENU_ITEMS and item not in deduped:
                deduped.append(item)
        clean[role] = deduped
    # a role omitted from the payload is reset to its default row rather than left dangling
    for role in ROLE_KEYS:
        clean.setdefault(role, list(DEFAULT_ROLE_MENUS.get(role, [])))
    setting = db.get(AppSetting, ROLE_MENUS_KEY)
    if not setting:
        setting = AppSetting(key=ROLE_MENUS_KEY, value="{}")
        db.add(setting)
    setting.value = json.dumps(clean)
    db.commit()
    return RoleMenusResponse(roles=list(ROLE_KEYS), items=list(MENU_ITEMS), menus=clean)


@router.post("/settings/environments", response_model=OptionList)
def add_environment(payload: OptionCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "environments")
    if payload.value not in values:
        values.append(payload.value)
        _save_setting_list(db, "environments", values)
    return get_options(_, db)


@router.post("/settings/server-types", response_model=OptionList)
def add_server_type(payload: OptionCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "server_types")
    if payload.value not in values:
        values.append(payload.value)
        _save_setting_list(db, "server_types", values)
    return get_options(_, db)


@router.post("/settings/application-types", response_model=OptionList)
def add_application_type(payload: OptionCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "application_types")
    if payload.value not in values:
        values.append(payload.value)
        _save_setting_list(db, "application_types", values)
    return get_options(_, db)


@router.delete("/settings/environments/{value}", response_model=OptionList)
def remove_environment(value: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "environments")
    if value in values:
        values.remove(value)
        _save_setting_list(db, "environments", values)
    return get_options(_, db)


@router.delete("/settings/server-types/{value}", response_model=OptionList)
def remove_server_type(value: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "server_types")
    if value in values:
        values.remove(value)
        _save_setting_list(db, "server_types", values)
    return get_options(_, db)


@router.delete("/settings/application-types/{value}", response_model=OptionList)
def remove_application_type(value: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> OptionList:
    values = _setting_list(db, "application_types")
    if value in values:
        values.remove(value)
        _save_setting_list(db, "application_types", values)
    return get_options(_, db)


# --- Vault secrets backend (feature/vault-secrets) -----------------------------------------
#
# Admin-only configuration for using HashiCorp Vault as the store for SSH credentials. All three
# endpoints are require_admin_not_guest: a desktop guest operates the local machine's DB but is
# not an account and must not touch a shared secrets backend. The token is sensitive, so GET
# never echoes it back (it returns token_set instead) and PUT only overwrites it when a non-empty
# token is supplied. When Vault stays disabled (the default) credentials keep using local Fernet.


def _vault_config_read(cfg) -> VaultConfigRead:
    return VaultConfigRead(
        enabled=cfg.enabled,
        address=cfg.address,
        kv_mount=cfg.kv_mount,
        path_prefix=cfg.path_prefix,
        token_set=cfg.token_set,
    )


@router.get("/settings/vault", response_model=VaultConfigRead)
def get_vault_settings(_: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> VaultConfigRead:
    return _vault_config_read(get_vault_config(db))


@router.put("/settings/vault", response_model=VaultConfigRead)
def update_vault_settings(payload: VaultConfigWrite, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> VaultConfigRead:
    cfg = save_vault_config(
        db,
        enabled=payload.enabled,
        address=payload.address,
        kv_mount=payload.kv_mount,
        path_prefix=payload.path_prefix,
        # empty string -> keep the stored token; save_vault_config treats falsy as "unchanged"
        token=payload.token,
    )
    return _vault_config_read(cfg)


@router.post("/settings/vault/test", response_model=VaultTestResult)
def test_vault_settings(payload: VaultConfigWrite, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> VaultTestResult:
    # Test the config as typed in the form without persisting it: use the posted fields, but fall
    # back to the stored token when the form's token box was left blank (the UI never re-shows it).
    stored = get_vault_config(db)
    from app.services.secrets import VaultConfig

    candidate = VaultConfig(
        enabled=payload.enabled,
        address=(payload.address or "").strip(),
        token=payload.token or stored.token,
        kv_mount=(payload.kv_mount or "").strip() or "secret",
        path_prefix=((payload.path_prefix or "").strip().strip("/") or "inframonitor"),
    )
    ok, message = test_vault(candidate)
    return VaultTestResult(ok=ok, message=message)


# --- app database backend switch (feature/app-db-backend) ----------------------------------
#
# Switch Infra Monitor's OWN backing database from the built-in SQLite to an external PostgreSQL
# (primary) or MySQL. The engine is a process-level singleton built at import, so the switch cannot
# be applied live: migrate COPIES all data into the target and writes an override file that the
# NEXT start reads. All four endpoints are require_admin_not_guest -- a desktop guest runs against
# the local machine's DB and must not repoint the whole install at a shared database. The migrate
# copy is never a move: on any failure the override file is not written and the source SQLite is
# left intact.


def _app_db_target_url(payload: AppDbConnectionRequest) -> str:
    """Resolve the request into a SQLAlchemy URL: an explicit `url`, else the discrete fields.

    Raises HTTP 400 on a bad engine or missing host, so the driver layer never sees garbage.
    """
    if payload.url.strip():
        return payload.url.strip()
    engine_kind = (payload.engine or "").strip().lower()
    if engine_kind not in ("postgres", "mysql"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="engine must be 'postgres' or 'mysql'")
    if not payload.host.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="host is required")
    try:
        return db_backend.build_url(
            engine_kind,
            payload.host.strip(),
            payload.port,
            payload.username,
            payload.password,
            payload.database.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.get("/settings/database", response_model=AppDbConfigRead)
def get_app_database(_: dict = Depends(require_admin)) -> AppDbConfigRead:
    cfg = db_backend.current_config()
    return AppDbConfigRead(backend=cfg["backend"], url_masked=cfg["url_masked"], is_override=cfg["is_override"])


@router.post("/settings/database/test", response_model=AppDbTestResult)
def test_app_database(payload: AppDbConnectionRequest, _: dict = Depends(require_admin)) -> AppDbTestResult:
    target_url = _app_db_target_url(payload)
    ok, message = db_backend.test_url(target_url)
    return AppDbTestResult(ok=ok, message=message)


@router.post("/settings/database/migrate", response_model=AppDbMigrateResult)
def migrate_app_database(payload: AppDbConnectionRequest, _: dict = Depends(require_admin)) -> AppDbMigrateResult:
    target_url = _app_db_target_url(payload)
    # Guard: writing the override needs a SQLite base configuration to anchor the file. Fail here
    # with a clear message rather than after a full (wasted) copy that could not be persisted.
    if db_backend.override_path() is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This install is not running on the default SQLite backend, so the database cannot be switched from here.",
        )
    result = db_backend.migrate_to(target_url)
    if not result["ok"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result["message"])
    # only persist the override once the copy has fully succeeded
    db_backend.write_override(target_url)
    return AppDbMigrateResult(ok=True, message=result["message"], tables=result["tables"], restart_required=True)


@router.post("/settings/database/use", response_model=AppDbMigrateResult)
def use_app_database(payload: AppDbConnectionRequest, _: dict = Depends(require_admin)) -> AppDbMigrateResult:
    # Point at an EXISTING database without copying anything -- for a target that already holds
    # Infra Monitor's data (a prior migration, or a shared/pre-provisioned DB). Startup will
    # create_all/seed against it, so an empty target is initialised and a populated one is kept.
    target_url = _app_db_target_url(payload)
    if db_backend.override_path() is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This install is not running on the default SQLite backend, so the database cannot be switched from here.",
        )
    result = db_backend.use_existing(target_url)
    if not result["ok"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result["message"])
    return AppDbMigrateResult(ok=True, message=result["message"], tables={}, restart_required=True)


@router.post("/settings/database/reset", response_model=AppDbMigrateResult)
def reset_app_database(_: dict = Depends(require_admin)) -> AppDbMigrateResult:
    removed = db_backend.clear_override()
    message = (
        "Reverted to the built-in SQLite database. Restart the app to apply."
        if removed
        else "Already using the built-in SQLite database; nothing to reset."
    )
    return AppDbMigrateResult(ok=True, message=message, tables={}, restart_required=removed)


# --- folders (EXPERIMENTAL: feature/server-folders) ----------------------------------------
#
# A folder is an optional, flat grouping over servers -- "BH", "MH", "EMS" and the like. There
# is no hierarchy and a server belongs to at most one folder. Mutations are admin-only; reads
# follow the same require_user rule the inventory does. The client only ever sees a folder's
# public_id, never its autoincrement primary key.


def _folder_or_404(db: Session, public_id: str) -> Folder:
    folder = db.scalar(select(Folder).where(Folder.public_id == public_id))
    if not folder:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Folder not found")
    return folder


def _folder_counts(db: Session) -> dict[int, int]:
    # one grouped query for every folder's membership, rather than reading folder.servers per
    # folder and calling len() -- that would lazy-load every member row just to count it
    rows = db.execute(
        select(Server.folder_id, func.count(Server.id))
        .where(Server.folder_id.is_not(None))
        .group_by(Server.folder_id)
    ).all()
    return {folder_id: count for folder_id, count in rows}


@router.get("/folders", response_model=list[FolderRead])
def list_folders(_: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[FolderRead]:
    counts = _folder_counts(db)
    # ordered case-insensitively so "BH" and "bh"-style names sort where a reader expects, not by
    # ASCII where uppercase sorts before lowercase
    folders = db.scalars(select(Folder).order_by(func.lower(Folder.name))).all()
    return [FolderRead(id=folder.public_id, name=folder.name, server_count=counts.get(folder.id, 0)) for folder in folders]


@router.post("/folders", response_model=FolderRead, status_code=status.HTTP_201_CREATED)
def create_folder(payload: FolderCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> FolderRead:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Folder name is required")
    # case-insensitive duplicate guard: "BH" and "bh" are the same folder to a human, and the
    # column's UNIQUE constraint alone is case-sensitive on SQLite
    if db.scalar(select(Folder).where(func.lower(Folder.name) == name.lower())):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A folder with that name already exists")
    # public_id is set here and only here: folders are born through this route, so there is no
    # legacy row that could exist without one, and no separate backfill pass is needed
    folder = Folder(name=name, public_id=str(uuid.uuid4()))
    db.add(folder)
    db.commit()
    db.refresh(folder)
    return FolderRead(id=folder.public_id, name=folder.name, server_count=0)


@router.patch("/folders/{public_id}", response_model=FolderRead)
def rename_folder(public_id: str, payload: FolderCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> FolderRead:
    folder = _folder_or_404(db, public_id)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Folder name is required")
    # a folder may keep (or re-case) its own name; only a clash with a DIFFERENT folder is a 409
    clash = db.scalar(select(Folder).where(func.lower(Folder.name) == name.lower(), Folder.id != folder.id))
    if clash:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A folder with that name already exists")
    folder.name = name
    db.commit()
    db.refresh(folder)
    counts = _folder_counts(db)
    return FolderRead(id=folder.public_id, name=folder.name, server_count=counts.get(folder.id, 0))


@router.delete("/folders/{public_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_folder(public_id: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> None:
    folder = _folder_or_404(db, public_id)
    # Unassign members FIRST, explicitly. There is deliberately no ON DELETE cascade on the FK, so
    # with SQLite's PRAGMA foreign_keys=ON a delete leaving servers pointed at this folder would
    # raise an IntegrityError. Servers must outlive their folder -- they just fall back to the
    # "Unassigned" group.
    db.execute(update(Server).where(Server.folder_id == folder.id).values(folder_id=None))
    db.delete(folder)
    db.commit()


@router.put("/servers/{server_id}/folder", response_model=ServerRead)
def assign_server_folder(server_id: str, payload: ServerFolderUpdate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ServerRead:
    server = _server_or_404(db, server_id)
    target = (payload.folder_id or "").strip()
    if not target:
        # empty string or null means "remove from whatever folder it was in"
        server.folder_id = None
    else:
        server.folder_id = _folder_or_404(db, target).id
    db.commit()
    db.refresh(server)
    return to_read(server)


@router.get("/servers", response_model=list[ServerRead])
def list_servers(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[ServerRead]:
    if _sees_all_servers(claims):
        return InventoryService(db).list_servers()
    user = _current_user(db, claims)
    return InventoryService(db).list_servers(_accessible_server_ids(db, user))


# --- inventory export -----------------------------------------------------------------------
#
# One .xlsx of everything the caller can see, for the reporting and hand-over asks that a screen
# cannot serve. It is deliberately a NON-SENSITIVE export: the SSH username is included because it
# is operational detail, but passwords and private keys never are -- only a yes/no "Credentials"
# column. Non-admins get exactly the servers they have access to, same rule as GET /servers.

_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Hostname", "hostname"),
    ("Alias", "alias"),
    ("IP Address", "ip_address"),
    ("SSH Port", "ssh_port"),
    ("Username", "username"),
    ("Group", "__group__"),
    ("Environment", "environment"),
    ("Server Type", "server_type"),
    ("Status", "status"),
    ("Operating System", "__os__"),
    ("Kernel", "kernel"),
    ("Architecture", "architecture"),
    ("CPU", "cpu"),
    ("CPU %", "__cpu_percent__"),
    ("RAM (MB)", "ram_mb"),
    ("RAM Used (MB)", "ram_used_mb"),
    ("Disk (GB)", "disk_gb"),
    ("Uptime (hours)", "__uptime_hours__"),
    ("Health Score", "health_score"),
    ("Docker", "docker_version"),
    ("Tags", "__tags__"),
    ("Business Owner", "business_owner"),
    ("Support Contact", "support_contact"),
    ("Credentials", "__has_credentials__"),
    ("Metrics Enabled", "__metrics__"),
    ("Last Checked", "__checked__"),
]


def _export_cell(server: ServerRead, key: str, group_name: str) -> object:
    if key == "__group__":
        return group_name or "Unassigned"
    if key == "__os__":
        return " ".join(part for part in [server.os_distro or server.operating_system, server.os_version] if part)
    if key == "__tags__":
        return ", ".join(server.tags)
    # cpu_percent is -1 when the host has never been sampled; an empty cell says "unknown" in a
    # spreadsheet, where -1 would be read as a real reading.
    if key == "__cpu_percent__":
        return "" if server.cpu_percent < 0 else server.cpu_percent
    if key == "__uptime_hours__":
        return round(server.uptime_seconds / 3600, 1) if server.uptime_seconds else ""
    if key == "__has_credentials__":
        return "Yes" if server.has_credentials else "No"
    if key == "__metrics__":
        return "Yes" if server.metrics_enabled else "No"
    if key == "__checked__":
        value = getattr(server, "vitals_checked_at", None) or getattr(server, "last_health_check", None)
        return str(value) if value else ""
    return getattr(server, key, "")


@router.get("/servers/export.xlsx")
def export_servers_xlsx(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - only when the optional wheel is missing
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export needs the openpyxl package. Install backend requirements and restart.",
        ) from exc

    if _sees_all_servers(claims):
        servers = InventoryService(db).list_servers()
    else:
        servers = InventoryService(db).list_servers(_accessible_server_ids(db, _current_user(db, claims)))
    # One id->name lookup for the Group column, rather than a folder query per row.
    group_names = {
        public_id: name
        for public_id, name in db.execute(select(Server.public_id, Folder.name).join(Folder, Server.folder_id == Folder.id)).all()
    }

    book = Workbook()
    sheet = book.active
    sheet.title = "Server Inventory"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3B57")
    for column, (label, _key) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row_index, server in enumerate(servers, start=2):
        group_name = group_names.get(server.id, "")
        for column, (_label, key) in enumerate(_EXPORT_COLUMNS, start=1):
            sheet.cell(row=row_index, column=column, value=_export_cell(server, key, group_name))
    # Freeze the header and size each column to its widest cell (capped), so the sheet is readable
    # the moment it opens rather than after a manual column-drag.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_EXPORT_COLUMNS))}{max(len(servers) + 1, 1)}"
    for column, (label, _key) in enumerate(_EXPORT_COLUMNS, start=1):
        widest = max([len(label)] + [len(str(sheet.cell(row=r, column=column).value or "")) for r in range(2, len(servers) + 2)])
        sheet.column_dimensions[get_column_letter(column)].width = min(max(widest + 2, 10), 42)

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="server-inventory-{stamp}.xlsx"'},
    )


def _validate_address_or_400(value: str) -> None:
    try:
        validate_host_address(value)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def _reject_duplicate(db: Session, *, hostname: str = "", ip_address: str = "", folder_id: int | None = None, exclude_id: int | None = None) -> None:
    # hostname/ip are not DB-unique, but CSV import refuses collisions, so a rename matches that
    # stance: changing a server's hostname or IP to one another server already holds is a 409.
    # exclude_id excludes the server being edited, so a no-op save is not a self-collision.
    #
    # Hostname uniqueness is scoped to the GROUP (folder_id): the same hostname may live in
    # different groups (BH and MH can each have a "dev" host), so a clash is only a clash within the
    # same folder_id. NULL/unassigned is its own bucket. IP addresses stay globally unique.
    if hostname:
        folder_predicate = Server.folder_id.is_(None) if folder_id is None else Server.folder_id == folder_id
        clash = db.scalar(select(Server).where(func.lower(Server.hostname) == hostname.strip().lower(), folder_predicate))
        if clash and clash.id != exclude_id:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"A server named {hostname} already exists in this group")
    if ip_address:
        clash = db.scalar(select(Server).where(func.lower(Server.ip_address) == ip_address.strip().lower()))
        if clash and clash.id != exclude_id:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Another server already uses IP address {ip_address}")


@router.post("/servers", response_model=ServerRead, status_code=status.HTTP_201_CREATED)
def create_server(payload: ServerCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ServerRead:
    _validate_address_or_400(payload.ip_address)
    # resolve the optional group (folder public_id) and enforce hostname uniqueness WITHIN it:
    # the same hostname may exist in different groups, so a clash is only rejected inside one folder.
    folder_id = _folder_or_404(db, payload.folder_id.strip()).id if (payload.folder_id or "").strip() else None
    _reject_duplicate(db, hostname=(payload.hostname or "").strip(), folder_id=folder_id)
    created = InventoryService(db).create_server(payload, folder_id=folder_id)
    server = db.scalar(select(Server).where(Server.public_id == created.id))
    if server and (payload.password or payload.private_key):
        _save_credentials(db, server,CredentialPayload(password=payload.password, private_key=payload.private_key))
        db.commit()
    if payload.password or payload.private_key:
        if server:
            try:
                _apply_discovery(db, server, _credentials_for(db, server,CredentialPayload()))
            except SshOperationError as exc:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Server saved, but discovery failed: {exc}") from exc
            db.refresh(server)
            return to_read(server)
    return created


@router.post("/servers/import", response_model=ServerImportResult)
def import_servers_api(payload: ServerImportRequest, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ServerImportResult:
    try:
        return import_servers(db, payload.csv_text, payload.dry_run)
    except CsvImportError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.get("/servers/{server_id}", response_model=ServerRead)
def get_server(server_id: str, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ServerRead:
    server = _server_or_404(db, server_id, claims)
    return to_read(server)


def _save_credentials(db: Session, server: Server, credentials: CredentialPayload) -> None:
    # Routes through the secrets abstraction: Vault when enabled+reachable, else the local
    # Fernet path (the default, unchanged). A Vault outage raises VaultError before any DB
    # column is touched, so it surfaces as a clear error instead of corrupting stored data.
    try:
        store_credentials(db, server, credentials.password, credentials.private_key)
    except VaultError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Vault error: {exc}") from exc


def _credentials_for(db: Session, server: Server, provided: CredentialPayload | None) -> CredentialPayload:
    if provided and (provided.password or provided.private_key):
        return provided
    try:
        password, private_key = load_credentials(db, server)
    except VaultError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Vault error: {exc}") from exc
    if not password and not private_key:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No stored credentials for this server. Ask an admin to save credentials.")
    return CredentialPayload(password=password, private_key=private_key, tail=provided.tail if provided else 200)


def _apply_discovery(db: Session, server: Server, credentials: CredentialPayload) -> None:
    data = discover_host(server, credentials)
    server.operating_system = data.get("OS", server.operating_system)
    server.kernel = data.get("KERNEL", server.kernel)
    server.architecture = data.get("ARCH", server.architecture)
    server.cpu = data.get("CPU", server.cpu)
    server.ram_mb = int(data.get("RAM_MB") or 0)
    server.disk_gb = int(data.get("DISK_GB") or 0)
    server.docker_version = data.get("DOCKER", "")
    server.podman_version = data.get("PODMAN", "")
    server.discovered_services_json = data.get("SERVICES_JSON", "[]")
    server.storage_json = data.get("STORAGE_JSON", "[]")
    server.database_logs_json = data.get("DB_LOGS_JSON", "[]")
    server.os_family = data.get("OS_FAMILY") or server.os_family
    server.os_distro = data.get("OS_DISTRO") or data.get("OS_ID") or server.os_distro
    server.os_version = data.get("OS_VERSION_ID") or server.os_version
    server.package_manager = data.get("PKG_MANAGER") or server.package_manager
    server.tomcat_json = data.get("TOMCAT_JSON") or server.tomcat_json
    services = ["ssh"]
    if server.docker_version:
        services.append("docker")
    if server.podman_version:
        services.append("podman")
    server.installed_services = ",".join(services)
    server.last_discovery = datetime.now(timezone.utc)
    server.last_seen = server.last_discovery
    # discovery already holds a working connection, so fold the vitals sample in rather
    # than making the caller run a second probe
    try:
        _apply_vitals(server, probe_vitals(server, credentials))
    except SshOperationError:
        pass
    # scored after the vitals fold-in so memory and load are available to it
    _apply_health(server)
    db.commit()


# --- health scoring ------------------------------------------------------------------------
#
# health_score used to be assigned a literal 100 by discovery, so it could never report a
# problem and the dashboard's "Health Score" tile was decoration. It is now derived from
# facts already collected: filesystem usage, memory usage, load per core, and any systemd
# unit discovery found in a failed state.
#
# A server with nothing measured yet stays `unknown` with a score of -1 rather than 0. Zero
# is indistinguishable from "measured and terrible", and averaging never-probed servers in
# as 0 is what made a healthy fleet read 50%.
HEALTH_UNMEASURED = -1


def _worst_disk_percent(server: Server) -> int:
    worst = -1
    for row in _json_list(server.storage_json):
        mount = str(row.get("mount") or "")
        kind = str(row.get("type") or "").lower()
        # skip pseudo-filesystems, but never the root, which is overlay on a container host
        if mount != "/" and kind in {"tmpfs", "devtmpfs", "squashfs", "overlay", "proc", "sysfs"}:
            continue
        raw = str(row.get("used_percent_num") or "").strip() or str(row.get("used_percent") or "").rstrip("%")
        try:
            worst = max(worst, int(float(raw)))
        except (TypeError, ValueError):
            continue
    return worst


def _failed_units(server: Server) -> int:
    return sum(
        1
        for row in _json_list(server.discovered_services_json)
        if str(row.get("source")) == "systemd" and str(row.get("status")).lower() == "failed"
    )


def _health_for(server: Server) -> tuple[int, ServerStatus]:
    penalties: list[int] = []

    disk = _worst_disk_percent(server)
    if disk >= 95:
        penalties.append(40)
    elif disk >= 90:
        penalties.append(25)
    elif disk >= 80:
        penalties.append(10)

    if server.ram_mb > 0 and server.ram_used_mb > 0:
        used = (server.ram_used_mb / server.ram_mb) * 100
        if used >= 95:
            penalties.append(20)
        elif used >= 90:
            penalties.append(10)
        elif used >= 80:
            penalties.append(5)

    # load is only meaningful per core, and `cpu` holds the core count from nproc
    try:
        cores = max(1, int(str(server.cpu).strip() or 1))
        load1 = float(str(server.load_average).split()[0])
        per_core = load1 / cores
        if per_core >= 2:
            penalties.append(20)
        elif per_core >= 1.5:
            penalties.append(10)
        elif per_core >= 1:
            penalties.append(5)
    except (TypeError, ValueError, IndexError):
        pass

    failed = _failed_units(server)
    if failed:
        penalties.append(min(30, failed * 10))

    measured = disk >= 0 or (server.ram_mb > 0 and server.ram_used_mb > 0) or bool(server.load_average)
    if not measured:
        return HEALTH_UNMEASURED, ServerStatus.unknown

    score = max(0, 100 - sum(penalties))
    if score >= 85:
        status = ServerStatus.healthy
    elif score >= 60:
        status = ServerStatus.warning
    else:
        status = ServerStatus.critical
    return score, status


def _apply_health(server: Server) -> None:
    score, status = _health_for(server)
    if score == HEALTH_UNMEASURED:
        return
    server.health_score = score
    server.status = status
    server.last_health_check = datetime.now(timezone.utc)


def _apply_vitals(server: Server, facts: dict) -> None:
    server.uptime_seconds = _as_int(facts.get("UPTIME_SECONDS"), server.uptime_seconds)
    server.load_average = facts.get("LOAD_AVERAGE") or server.load_average
    server.cpu_percent = _as_int(facts.get("CPU_PERCENT"), -1)
    server.ram_used_mb = _as_int(facts.get("RAM_USED_MB"), server.ram_used_mb)
    server.process_count = _as_int(facts.get("PROCESS_COUNT"), server.process_count)
    total = _as_int(facts.get("RAM_TOTAL_MB"))
    if total:
        server.ram_mb = total
    server.vitals_checked_at = datetime.now(timezone.utc)
    # A successful probe is itself evidence the host is reachable, so a server imported by CSV
    # and never discovered stops sitting at "unknown" with a zero score.
    server.last_seen = server.vitals_checked_at
    _apply_health(server)


def _as_int(value: object, fallback: int = 0) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return fallback


_JOURNAL_UNIT = re.compile(r"^[A-Za-z0-9@._:-]{1,128}$")
_UNKNOWN_LOG_PATH = "Log path is not a discovered log source for this server. Run discovery first."
_SUDO_REJECTED_MARKERS = ("incorrect password", "sorry, try again", "authentication failure", "authentication failed")
_SUDO_REQUIRED_MARKERS = ("a password is required", "no tty present", "askpass", "sudo: a terminal is required")


def _json_list(value: str) -> list[dict]:
    try:
        parsed = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    return [item for item in parsed if isinstance(item, dict)] if isinstance(parsed, list) else []


def _known_log_paths(server: Server) -> set[str]:
    paths: set[str] = set()
    for entry in _json_list(server.database_logs_json):
        path = str(entry.get("path") or "").strip()
        if path:
            paths.add(path)
    for instance in _json_list(server.tomcat_json):
        log_files = instance.get("log_files")
        if not isinstance(log_files, list):
            continue
        for log_file in log_files:
            path = str(log_file.get("path") or "").strip() if isinstance(log_file, dict) else ""
            if path:
                paths.add(path)
    return paths


def _validated_log_target(server: Server, source: str, name_or_path: str) -> str:
    target = name_or_path.strip()
    if source == "journal":
        if not _JOURNAL_UNIT.fullmatch(target):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Journal unit name contains unsupported characters")
        return target
    if target not in _known_log_paths(server):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_UNKNOWN_LOG_PATH)
    return target


def _sudo_prompt(server: Server) -> PrivilegedOperationResult:
    host = server.hostname or server.ip_address
    return PrivilegedOperationResult(ok=False, needs_sudo_password=True, message=f"Sudo password required for {server.username}@{host}")


def _is_sudo_signal(exc: SshOperationError, password_supplied: bool) -> bool:
    lowered = str(exc).lower()
    markers = _SUDO_REJECTED_MARKERS + _SUDO_REQUIRED_MARKERS if password_supplied else _SUDO_REQUIRED_MARKERS
    return any(marker in lowered for marker in markers)


def _privileged_failure(server: Server, exc: SshOperationError, password_supplied: bool) -> PrivilegedOperationResult:
    if _is_sudo_signal(exc, password_supplied):
        if password_supplied:
            return PrivilegedOperationResult(ok=False, needs_sudo_password=True, message="Sudo authentication failed")
        return _sudo_prompt(server)
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.delete("/servers/{server_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_server(server_id: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> None:
    server = _server_or_404(db, server_id)
    db.delete(server)
    db.commit()


_EDITABLE_TEXT_FIELDS = ("hostname", "alias", "ip_address", "username", "environment", "server_type", "business_owner", "support_contact", "os_kind")


@router.patch("/servers/{server_id}", response_model=ServerRead)
def update_server(server_id: str, payload: ServerUpdate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ServerRead:
    server = _server_or_404(db, server_id)
    # exclude_unset: only touch fields the client actually sent, so an edit dialog that submits
    # a subset never blanks the rest.
    data = payload.model_dump(exclude_unset=True)

    if "ip_address" in data:
        ip = (data["ip_address"] or "").strip()
        if not ip:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="IP address / hostname cannot be empty")
        _validate_address_or_400(ip)
    if "hostname" in data and not (data["hostname"] or "").strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Hostname cannot be empty")
    if "username" in data and not (data["username"] or "").strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SSH user cannot be empty")
    if "ssh_port" in data and data["ssh_port"] is not None:
        port = int(data["ssh_port"])
        if port < 1 or port > 65535:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SSH port must be between 1 and 65535")
    _reject_duplicate(
        db,
        hostname=data["hostname"].strip() if data.get("hostname") else "",
        ip_address=data["ip_address"].strip() if data.get("ip_address") else "",
        # a hostname clash is scoped to the server's current group; a rename doesn't move groups
        folder_id=server.folder_id,
        exclude_id=server.id,
    )

    for field in _EDITABLE_TEXT_FIELDS:
        if field in data and data[field] is not None:
            setattr(server, field, data[field].strip())
    if "ssh_port" in data and data["ssh_port"] is not None:
        server.ssh_port = int(data["ssh_port"])
    if "tags" in data and data["tags"] is not None:
        server.tags = ",".join(sorted({tag.strip() for tag in data["tags"] if tag.strip()}))

    db.commit()
    db.refresh(server)
    return to_read(server)


@router.put("/servers/{server_id}/credentials", response_model=ConnectionResult)
def update_credentials(server_id: str, credentials: CredentialPayload, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id)
    _save_credentials(db, server,credentials)
    db.commit()
    return ConnectionResult(ok=True, message="Credentials saved encrypted")


@router.post("/servers/{server_id}/test-connection", response_model=ConnectionResult)
def test_connection(server_id: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id, claims)
    try:
        # Use the supplied credentials when the caller typed some (validate before saving);
        # otherwise fall back to the stored ones so a plain "Test connection" checks the
        # existing server without re-entering the password.
        creds = credentials if (credentials.password or credentials.private_key) else _credentials_for(db, server, credentials)
        output = run_command(server, creds, "hostname")
        return ConnectionResult(ok=True, message=f"Connected to {output.strip()}")
    except SshOperationError as exc:
        return ConnectionResult(ok=False, message=str(exc))


@router.post("/servers/{server_id}/discover", response_model=ConnectionResult)
def discover(server_id: str, credentials: CredentialPayload, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id)
    try:
        if credentials.password or credentials.private_key:
            _save_credentials(db, server,credentials)
            db.commit()
        _apply_discovery(db, server, _credentials_for(db, server,credentials))
        return ConnectionResult(ok=True, message="Discovery completed")
    except SshOperationError as exc:
        return ConnectionResult(ok=False, message=str(exc))


@router.post("/servers/{server_id}/containers/{runtime}", response_model=list[ContainerRead])
def containers(server_id: str, runtime: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[ContainerRead]:
    if runtime not in {"docker", "podman"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Runtime must be docker or podman")
    server = _server_or_404(db, server_id, claims)
    try:
        return list_containers_with_ports(server, _credentials_for(db, server,credentials), runtime)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/images/{runtime}", response_model=list[ImageRead])
def images(server_id: str, runtime: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[ImageRead]:
    if runtime not in {"docker", "podman"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Runtime must be docker or podman")
    server = _server_or_404(db, server_id, claims)
    try:
        return list_images(server, _credentials_for(db, server, credentials), runtime)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/logs/{runtime}/{container}", response_model=LogResponse)
def logs(server_id: str, runtime: str, container: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> LogResponse:
    if runtime not in {"docker", "podman"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Runtime must be docker or podman")
    server = _server_or_404(db, server_id, claims)
    try:
        effective = _credentials_for(db, server,credentials)
        return LogResponse(runtime=runtime, container=container, lines=container_logs(server, effective, runtime, container, effective.tail))
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/containers/{runtime}/{container}/env", response_model=LogResponse)
def container_env(server_id: str, runtime: str, container: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> LogResponse:
    if runtime not in {"docker", "podman"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Runtime must be docker or podman")
    server = _server_or_404(db, server_id, claims)
    try:
        lines = container_env_file(server, _credentials_for(db, server,credentials), runtime, container)
        return LogResponse(runtime=runtime, container=container, lines=lines)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/logs/service", response_model=LogResponse)
def service_log(server_id: str, payload: ServiceLogRequest, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> LogResponse:
    server = _server_or_404(db, server_id, claims)
    target = _validated_log_target(server, payload.source, payload.name_or_path)
    try:
        effective = _credentials_for(db, server,CredentialPayload())
        return LogResponse(runtime=payload.source, container=target, lines=service_logs(server, effective, payload.source, target, payload.tail))
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/tomcat", response_model=list[TomcatInstance])
def tomcat_instances(server_id: str, credentials: CredentialPayload, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[TomcatInstance]:
    server = _server_or_404(db, server_id, claims)
    try:
        instances = discover_tomcat(server, _credentials_for(db, server,credentials))
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    server.tomcat_json = json.dumps(instances)
    db.commit()
    return [TomcatInstance.model_validate(instance) for instance in instances]


@router.post("/servers/{server_id}/tomcat/logs", response_model=LogResponse)
def tomcat_log(server_id: str, payload: TomcatLogRequest, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> LogResponse:
    server = _server_or_404(db, server_id, claims)
    target = _validated_log_target(server, "file", payload.log_file)
    try:
        effective = _credentials_for(db, server,CredentialPayload())
        return LogResponse(runtime="tomcat", container=target, lines=tomcat_logs(server, effective, target, payload.tail))
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/tomcat/action", response_model=PrivilegedOperationResult)
def tomcat_action_api(server_id: str, payload: TomcatActionRequest, claims: dict = Depends(require_admin), db: Session = Depends(get_db)) -> PrivilegedOperationResult:
    server = _server_or_404(db, server_id, claims)
    action = payload.action.strip().lower() or "restart"
    if action not in {"restart", "start", "stop", "status"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Action must be restart, start, stop, or status")
    instance = payload.instance.strip()
    if not instance:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tomcat instance is required")
    try:
        output = tomcat_action(server, _credentials_for(db, server,CredentialPayload()), instance, action, payload.sudo_password)
    except SudoPasswordRequired:
        return _sudo_prompt(server)
    except SshOperationError as exc:
        return _privileged_failure(server, exc, bool(payload.sudo_password))
    return PrivilegedOperationResult(ok=True, message=f"Tomcat {instance} {action} completed", output=output)


_ZIP_MAGIC = b"PK\x03\x04"
# literal, not starlette.status: the constant was renamed between Starlette releases
# (HTTP_413_REQUEST_ENTITY_TOO_LARGE -> HTTP_413_CONTENT_TOO_LARGE)
_HTTP_413 = 413
_WAR_NAME_MAX = 255
_WAR_READ_CHUNK = 1024 * 1024
_DEFAULT_MAX_WAR_MB = 512


def _max_war_bytes() -> int:
    # MAX_WAR_MB comes from config.py; a missing or unusable value must still leave a cap in
    # force rather than accepting an unbounded upload
    try:
        limit_mb = int(getattr(get_settings(), "max_war_mb", _DEFAULT_MAX_WAR_MB))
    except (TypeError, ValueError):
        limit_mb = _DEFAULT_MAX_WAR_MB
    return max(1, limit_mb) * 1024 * 1024


def _validated_war_name(candidate: str) -> str:
    # This is the only place a client-supplied name reaches a remote path. The target directory
    # is derived from the discovered instance, never from the request, and the name must be a
    # bare basename so it cannot climb out of <catalina_base>/webapps.
    name = (candidate or "").strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A WAR file name is required")
    if len(name) > _WAR_NAME_MAX:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"WAR file name must be at most {_WAR_NAME_MAX} characters")
    if "/" in name or "\\" in name or ".." in name or any(ord(char) < 32 or ord(char) == 127 for char in name):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="WAR file name must be a bare file name with no directory separators, '..' or control characters",
        )
    if name.startswith("-") or name.startswith("."):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="WAR file name must not start with '-' or '.'")
    # a bare ".war" is already caught by the leading-dot rule above
    if not name.lower().endswith(".war"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .war files can be deployed")
    return name


def _find_tomcat_instance(instances: list[dict], name: str) -> dict | None:
    for item in instances:
        if item.get("name") == name or (item.get("unit") and item.get("unit") == name):
            return item
    return None


def _tomcat_instance_or_404(db: Session, server: Server, instance: str) -> dict:
    name = instance.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tomcat instance is required")
    match = _find_tomcat_instance(_json_list(server.tomcat_json), name)
    if match is None:
        # nothing persisted for this instance yet -- probe once, then it is a genuine 404
        try:
            instances = discover_tomcat(server, _credentials_for(db, server,CredentialPayload()))
        except SshOperationError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
        # only persist a probe that actually saw something: overwriting with an empty list would
        # throw away the log-file allowlist that _known_log_paths depends on
        if instances:
            server.tomcat_json = json.dumps(instances)
            db.commit()
        match = _find_tomcat_instance(instances, name)
    if match is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Tomcat instance not found: {name}")
    return match


def _webapps_dir(instance: dict) -> str:
    base = str(instance.get("catalina_base") or "").strip()
    if not base.startswith("/") or "\x00" in base:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Discovered instance has no absolute CATALINA_BASE, so its webapps directory is unknown. Run discovery again.",
        )
    return f"{base.rstrip('/')}/webapps"


def _read_capped_upload(upload: UploadFile, limit_bytes: int, detail: str = "") -> bytes:
    # Starlette has already spooled the multipart body (to a temp file past ~1 MB) before this
    # runs, so the cap bounds what we hold in memory and hand to SFTP, not that spool file.
    too_large = detail or _war_too_large(limit_bytes)
    if getattr(upload, "size", None) and int(upload.size or 0) > limit_bytes:
        raise HTTPException(status_code=_HTTP_413, detail=too_large)
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = upload.file.read(_WAR_READ_CHUNK)
        if not chunk:
            break
        total += len(chunk)
        if total > limit_bytes:
            raise HTTPException(status_code=_HTTP_413, detail=too_large)
        chunks.append(chunk)
    return b"".join(chunks)


def _war_too_large(limit_bytes: int) -> str:
    if limit_bytes >= 1024 * 1024:
        return f"WAR is larger than the {limit_bytes // (1024 * 1024)} MB limit (MAX_WAR_MB)"
    return f"WAR is larger than the {limit_bytes} byte limit (MAX_WAR_MB)"


def _deploy_text(result: object, *keys: str) -> str:
    if not isinstance(result, dict):
        return ""
    for key in keys:
        value = result.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def _deploy_int(result: object, key: str, fallback: int) -> int:
    if isinstance(result, dict):
        try:
            return int(result.get(key))  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return fallback
    return fallback


@router.post("/servers/{server_id}/tomcat/deploy", response_model=WarDeployResult)
def tomcat_deploy_war(
    server_id: str,
    instance: str = Form(..., max_length=128),
    file: UploadFile = File(...),
    filename: str = Form("", max_length=_WAR_NAME_MAX),
    restart: bool = Form(False),
    sudo_password: str = Form("", max_length=512),
    claims: dict = Depends(require_admin),
    db: Session = Depends(get_db),
) -> WarDeployResult:
    server = _server_or_404(db, server_id, claims)
    # cheap, purely local checks first: never open an SSH connection for a request that
    # cannot possibly be valid
    war_name = _validated_war_name(filename or (file.filename or ""))
    target_instance = _tomcat_instance_or_404(db, server, instance)
    target_dir = _webapps_dir(target_instance)
    try:
        data = _read_capped_upload(file, _max_war_bytes())
    finally:
        file.file.close()
    if not data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded WAR is empty")
    if not data.startswith(_ZIP_MAGIC):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is not a WAR: it does not start with the ZIP magic PK\\x03\\x04",
        )

    credentials = _credentials_for(db, server,CredentialPayload())
    try:
        result = deploy_war(server, credentials, target_dir, war_name, data, sudo_password)
    except SudoPasswordRequired:
        prompt = _sudo_prompt(server)
        return WarDeployResult(ok=False, message=prompt.message, needs_sudo_password=True)
    except SshOperationError as exc:
        # raises 400 for anything that is not a sudo-password problem
        failure = _privileged_failure(server, exc, bool(sudo_password))
        return WarDeployResult(ok=False, message=failure.message, needs_sudo_password=failure.needs_sudo_password)

    target_path = _deploy_text(result, "target_path", "target") or f"{target_dir}/{war_name}"
    backup_path = _deploy_text(result, "backup_path", "backup")
    bytes_written = _deploy_int(result, "bytes_written", len(data))
    message = f"Deployed {war_name} to {target_dir}"
    if backup_path:
        message = f"{message}; previous file backed up to {backup_path}"
    deployed = WarDeployResult(
        ok=True,
        message=message,
        target_path=target_path,
        backup_path=backup_path,
        bytes_written=bytes_written,
    )
    if not restart:
        return deployed

    instance_name = str(target_instance.get("name") or instance.strip())
    try:
        tomcat_action(server, credentials, instance_name, "restart", sudo_password)
    except SudoPasswordRequired:
        # The WAR is on disk. Report success with restarted=False so the UI prompts for the
        # sudo password and retries the restart alone -- re-uploading would be wrong.
        return deployed.model_copy(
            update={
                "needs_sudo_password": True,
                "message": f"{message}. The restart needs a sudo password; retry the restart only, the WAR is already in place.",
            }
        )
    except SshOperationError as exc:
        # Same reasoning: a failed restart must not surface as a 4xx, or the UI reads it as a
        # failed deploy and uploads again.
        return deployed.model_copy(
            update={
                "needs_sudo_password": _is_sudo_signal(exc, bool(sudo_password)),
                "message": f"{message}, but the restart failed: {exc}",
            }
        )
    return deployed.model_copy(update={"restarted": True, "message": f"{message}; instance {instance_name} restarted"})


@router.post("/servers/{server_id}/container-restart", response_model=ConnectionResult)
def restart_container_api(server_id: str, payload: OperationRequest, claims: dict = Depends(require_admin_or_developer), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id, claims)
    try:
        result = restart_container(server, _credentials_for(db, server,CredentialPayload()), payload.runtime, payload.name)
        return ConnectionResult(ok=True, message=f"Restarted {result or payload.name}")
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


# Delete is destructive, so admin-only (stricter than restart's admin-or-developer).
@router.post("/servers/{server_id}/container-remove", response_model=ConnectionResult)
def remove_container_api(server_id: str, payload: OperationRequest, claims: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id, claims)
    try:
        remove_container(server, _credentials_for(db, server, CredentialPayload()), payload.runtime, payload.name)
        return ConnectionResult(ok=True, message=f"Removed container {payload.name}")
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/image-remove", response_model=ConnectionResult)
def remove_image_api(server_id: str, payload: OperationRequest, claims: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ConnectionResult:
    server = _server_or_404(db, server_id, claims)
    try:
        remove_image(server, _credentials_for(db, server, CredentialPayload()), payload.runtime, payload.name)
        return ConnectionResult(ok=True, message=f"Removed image {payload.name}")
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.post("/servers/{server_id}/services/restart", response_model=PrivilegedOperationResult)
def restart_service_api(server_id: str, payload: ServiceRestartRequest, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> PrivilegedOperationResult:
    server = _server_or_404(db, server_id)
    try:
        result = restart_service(server, _credentials_for(db, server,CredentialPayload()), payload.name, payload.sudo_password)
    except SudoPasswordRequired:
        return _sudo_prompt(server)
    except SshOperationError as exc:
        return _privileged_failure(server, exc, bool(payload.sudo_password))
    return PrivilegedOperationResult(ok=True, message=f"Service {payload.name} status: {result}", output=result)


@router.get("/dashboard/summary", response_model=Summary)
def summary(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> Summary:
    if _sees_all_servers(claims):
        servers = db.scalars(select(Server)).all()
    else:
        ids = _accessible_server_ids(db, _current_user(db, claims))
        servers = db.scalars(select(Server).where(Server.id.in_(ids))).all() if ids else []
    counts = {status.value: 0 for status in ServerStatus}
    for server in servers:
        counts[server.status.value] += 1

    # Average only over servers that have actually been measured. Including never-probed
    # servers as 0 made a fleet of healthy hosts plus two fresh CSV imports read as 50%.
    scored = [server.health_score for server in servers if server.status is not ServerStatus.unknown]

    container_hosts = 0
    database_services = 0
    active_services = 0
    for server in servers:
        if server.docker_version or server.podman_version:
            container_hosts += 1
        for row in _json_list(server.discovered_services_json):
            if str(row.get("type")) == "database":
                database_services += 1
            if str(row.get("status")).lower() == "active":
                active_services += 1

    return Summary(
        server_count=len(servers),
        healthy_servers=counts["healthy"],
        warning_servers=counts["warning"],
        critical_servers=counts["critical"],
        offline_servers=counts["offline"],
        # was counted but shown nowhere, so healthy+warning+critical+offline silently came to
        # less than server_count and the dashboard's numbers did not reconcile
        unknown_servers=counts["unknown"],
        # These three were literal zeros. They are counts of what discovery actually found, so
        # they are honest but not what their old names imply: this is the number of hosts with
        # a container runtime, not a count of containers (which needs a live probe per host).
        total_containers=container_hosts,
        databases=database_services,
        running_services=active_services,
        average_health_score=round(sum(scored) / len(scored)) if scored else 0,
        measured_servers=len(scored),
    )


@router.get("/integrations", response_model=list[IntegrationStatus])
async def integrations(_: dict = Depends(require_user)) -> list[IntegrationStatus]:
    return await check_integrations()


# --- Alertmanager receiver ------------------------------------------------------------------
#
# The buffer below is IN-PROCESS AND LOST ON RESTART. It is also per-process, so with more than
# one uvicorn worker each worker sees only the alerts it received. It exists so a fired alert is
# visible in the UI instead of vanishing; it is NOT durable alert storage, and deliberately has
# no database table -- Alertmanager itself remains the record of what is firing.
_ALERT_BUFFER_CAPACITY = 100
_ALERT_BUFFER_NOTE = (
    "In-memory ring buffer of the last 100 alerts received by this process. "
    "Lost on restart and not shared between workers -- visibility only, not durable storage."
)
_ALERT_MAX_PER_REQUEST = 500
_ALERT_FIELD_MAX = 512
_alert_buffer: deque[AlertRecord] = deque(maxlen=_ALERT_BUFFER_CAPACITY)
_alert_buffer_lock = threading.Lock()


def _alert_text(value: object) -> str:
    if value is None or isinstance(value, (dict, list, bytes)):
        return ""
    return str(value).strip()[:_ALERT_FIELD_MAX]


def _alert_mapping(value: object) -> dict:
    return value if isinstance(value, dict) else {}


def _alert_record(raw: object, received_at: datetime) -> AlertRecord | None:
    # every field is optional and every container is type-checked: a webhook must never 500 on a
    # payload shape we did not expect (the old version called len() on whatever "alerts" was)
    if not isinstance(raw, dict):
        return None
    labels = _alert_mapping(raw.get("labels"))
    annotations = _alert_mapping(raw.get("annotations"))
    record = AlertRecord(
        alertname=_alert_text(labels.get("alertname")) or _alert_text(raw.get("alertname")),
        severity=_alert_text(labels.get("severity")),
        status=_alert_text(raw.get("status")),
        instance=_alert_text(labels.get("instance")) or _alert_text(labels.get("job")),
        summary=_alert_text(annotations.get("summary")) or _alert_text(annotations.get("description")),
        starts_at=_alert_text(raw.get("startsAt")) or _alert_text(raw.get("starts_at")),
        received_at=received_at,
    )
    if not (record.alertname or record.summary or record.instance or record.status):
        return None
    return record


@router.post("/alerts/webhook", response_model=AlertWebhookResult)
def alerts_webhook(payload: dict[str, Any] = Body(default_factory=dict)) -> AlertWebhookResult:
    # Unauthenticated on purpose: Alertmanager posts here server-to-server inside the compose
    # network and has no way to carry a JWT. The blast radius is bounded -- nothing is written to
    # the database, and the worst a reachable poster can do is push noise through a 100-entry
    # buffer. Do not extend this endpoint to touch persistent state.
    alerts = payload.get("alerts")
    if not isinstance(alerts, list):
        # a group-level payload with no usable alerts array is accepted and ignored rather than
        # rejected: Alertmanager retries failed deliveries and we do not want a retry storm
        return AlertWebhookResult(ok=True, received=0, stored=0, message="No alerts array in payload; nothing recorded")
    received_at = datetime.now(timezone.utc)
    records = [record for raw in alerts[:_ALERT_MAX_PER_REQUEST] if (record := _alert_record(raw, received_at))]
    if records:
        with _alert_buffer_lock:
            _alert_buffer.extend(records)
    return AlertWebhookResult(
        ok=True,
        received=len(alerts),
        stored=len(records),
        message=f"Buffered {len(records)} of {len(alerts)} alerts in memory (lost on restart)",
    )


@router.get("/alerts/recent", response_model=AlertBufferResponse)
def recent_alerts(_: dict = Depends(require_user)) -> AlertBufferResponse:
    with _alert_buffer_lock:
        items = list(_alert_buffer)
    items.reverse()  # newest first
    return AlertBufferResponse(
        alerts=items,
        count=len(items),
        capacity=_ALERT_BUFFER_CAPACITY,
        persistent=False,
        note=_ALERT_BUFFER_NOTE,
    )


# --- monitoring proxies (full profile only) -------------------------------------------------
#
# Server-side proxies to the compose-network Prometheus / Loki / Alertmanager so the app's own UI
# can render metrics, logs and alerts. require_user (a signed-in user OR the desktop guest) may
# read them. The upstream base URL is fixed from settings -- only query params/label names vary --
# so a caller can never point these at an arbitrary host. A MonitoringError (missing URL or an
# upstream/transport failure) maps to HTTP 502 with its clean message; the frontend calls
# /monitoring/enabled first and only shows the section when the relevant upstream is configured.


def _monitoring_502(exc: MonitoringError) -> HTTPException:
    return HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc))


def _monitoring_int(value: str, field: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=f"'{field}' must be an integer"
        ) from exc


@router.get("/monitoring/enabled", response_model=None)
def monitoring_enabled(_: dict = Depends(require_user)) -> dict:
    return monitoring.configured()


@router.get("/monitoring/prometheus/query", response_model=None)
async def monitoring_prometheus_query(query: str, _: dict = Depends(require_user)) -> dict:
    try:
        return await monitoring.prometheus_query(query)
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


@router.get("/monitoring/prometheus/query_range", response_model=None)
async def monitoring_prometheus_query_range(
    query: str, start: str, end: str, step: str, _: dict = Depends(require_user)
) -> dict:
    try:
        return await monitoring.prometheus_query_range(query, start, end, step)
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


@router.get("/monitoring/loki/query_range", response_model=None)
async def monitoring_loki_query_range(
    query: str,
    start: str,
    end: str,
    limit: str = "1000",
    direction: str = "backward",
    _: dict = Depends(require_user),
) -> dict:
    limit_n = _monitoring_int(limit, "limit")
    direction = direction if direction in ("backward", "forward") else "backward"
    try:
        return await monitoring.loki_query_range(query, start, end, limit_n, direction)
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


@router.get("/monitoring/loki/labels", response_model=None)
async def monitoring_loki_labels(_: dict = Depends(require_user)) -> dict:
    try:
        return await monitoring.loki_labels()
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


@router.get("/monitoring/loki/label/{name}/values", response_model=None)
async def monitoring_loki_label_values(name: str, _: dict = Depends(require_user)) -> dict:
    try:
        return await monitoring.loki_label_values(name)
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


@router.get("/monitoring/alerts", response_model=None)
async def monitoring_alerts(_: dict = Depends(require_user)) -> list:
    try:
        return await monitoring.alertmanager_alerts()
    except MonitoringError as exc:
        raise _monitoring_502(exc) from exc


# --- per-server monitoring ingestion --------------------------------------------------------
#
# Two capabilities, both opt-in per server and both mutated only through the endpoints here (never
# PATCH /servers): (a) install node_exporter over SSH so Prometheus scrapes the host's metrics, and
# (b) tail the server's logs into Loki. The install/uninstall run privileged commands on the real
# host and reuse the shared sudo-aware PrivilegedOperationResult flow; the http_sd endpoint is what
# Prometheus itself calls to discover which servers to scrape.


def _require_sd_token(authorization: str | None) -> None:
    """Gate the http_sd target list on settings.monitoring_sd_token.

    When the token is unset the endpoint is left open, so a dev stack without a configured token
    still works (documented on the setting). When set, Prometheus must present it as a bearer token
    in the Authorization header; anything else is a 401.
    """
    expected = (get_settings().monitoring_sd_token or "").strip()
    if not expected:
        return
    presented = (authorization or "").strip()
    prefix = "Bearer "
    token = presented[len(prefix):].strip() if presented.startswith(prefix) else ""
    if token != expected:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or missing bearer token",
            headers={"WWW-Authenticate": "Bearer"},
        )


@router.get("/monitoring/prometheus/targets", response_model=None)
def prometheus_targets(authorization: str | None = Header(default=None), db: Session = Depends(get_db)) -> list[dict]:
    """Prometheus http_sd target list: every metrics-enabled server as a scrape target.

    Authenticated by the shared MONITORING_SD_TOKEN bearer token (see _require_sd_token), NOT the
    normal user auth -- this is a machine endpoint Prometheus polls, not a UI route.
    """
    _require_sd_token(authorization)
    servers = db.scalars(select(Server).where(Server.metrics_enabled.is_(True))).all()
    targets: list[dict] = []
    for server in servers:
        targets.append(
            {
                "targets": [f"{server.ip_address}:{server.node_exporter_port}"],
                "labels": {
                    "server": server.hostname or server.ip_address,
                    "server_id": server.public_id,
                    "env": server.environment,
                    "job": "node",
                },
            }
        )
    return targets


@router.post("/servers/{server_id}/monitoring/install-metrics", response_model=PrivilegedOperationResult)
def install_metrics(server_id: str, payload: MonitoringInstallRequest = Body(default=MonitoringInstallRequest()), _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> PrivilegedOperationResult:
    server = _server_or_404(db, server_id)
    try:
        ok, message = node_exporter.install(server, _credentials_for(db, server, CredentialPayload()), payload.sudo_password)
    except SudoPasswordRequired:
        return _sudo_prompt(server)
    except NodeExporterError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    except SshOperationError as exc:
        return _privileged_failure(server, exc, bool(payload.sudo_password))
    server.metrics_enabled = True
    db.commit()
    return PrivilegedOperationResult(ok=ok, message=message)


@router.post("/servers/{server_id}/monitoring/uninstall-metrics", response_model=PrivilegedOperationResult)
def uninstall_metrics(server_id: str, payload: MonitoringInstallRequest = Body(default=MonitoringInstallRequest()), _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> PrivilegedOperationResult:
    server = _server_or_404(db, server_id)
    try:
        ok, message = node_exporter.uninstall(server, _credentials_for(db, server, CredentialPayload()), payload.sudo_password)
    except SudoPasswordRequired:
        return _sudo_prompt(server)
    except NodeExporterError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    except SshOperationError as exc:
        return _privileged_failure(server, exc, bool(payload.sudo_password))
    server.metrics_enabled = False
    db.commit()
    return PrivilegedOperationResult(ok=ok, message=message)


@router.put("/servers/{server_id}/monitoring/log-shipping", response_model=ServerMonitoringState)
def update_log_shipping(server_id: str, payload: LogShippingUpdate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> ServerMonitoringState:
    server = _server_or_404(db, server_id)
    sources: list[dict] = []
    for entry in payload.sources:
        source = (entry.source or "journal").strip() or "journal"
        name_or_path = (entry.name_or_path or "").strip()
        if not name_or_path:
            continue
        sources.append({"source": source, "name_or_path": name_or_path})
    server.log_shipping_enabled = bool(payload.enabled)
    server.log_sources_json = json.dumps(sources)
    db.commit()
    db.refresh(server)
    return ServerMonitoringState(
        metrics_enabled=bool(server.metrics_enabled),
        node_exporter_port=server.node_exporter_port,
        scraped=False,
        log_shipping_enabled=bool(server.log_shipping_enabled),
        log_sources=[LogSourceEntry(**item) for item in sources],
    )


@router.get("/servers/{server_id}/log-capabilities", response_model=None)
def server_log_capabilities(server_id: str, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> dict:
    """Detect which log-producing systems this host runs (nginx/docker/podman/kubernetes) and
    suggest log sources. One cheap SSH probe; returns all-false on any SSH failure (never 500s)."""
    server = _server_or_404(db, server_id, claims)
    return detect_log_capabilities(server, _credentials_for(db, server, CredentialPayload()))


# --- user-defined Prometheus alert rules (feature/custom-alerts) ---------------------------
# Written to a file shared with the Prometheus container; Prometheus is reloaded via /-/reload so
# changes apply without a restart. Reads are require_user; mutations are require_admin.


@router.get("/monitoring/alerts/rules", response_model=None)
def list_alert_rules(_: dict = Depends(require_user)) -> list[dict]:
    return alert_rules.list_rules()


@router.get("/monitoring/alerts/templates", response_model=None)
def alert_rule_templates(_: dict = Depends(require_user)) -> list[dict]:
    return alert_rules.TEMPLATES


@router.post("/monitoring/alerts/rules", response_model=None, status_code=status.HTTP_201_CREATED)
def add_alert_rule(payload: AlertRuleCreate, _: dict = Depends(require_admin)) -> dict:
    try:
        alert_rules.add_rule(
            alert_rules.AlertRule(
                name=payload.name.strip(),
                expr=payload.expr.strip(),
                for_duration=(payload.for_duration or "5m").strip(),
                severity=(payload.severity or "warning").strip(),
                summary=payload.summary.strip(),
                description=payload.description.strip(),
            )
        )
    except alert_rules.AlertRuleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return {"ok": True, "message": f"Alert rule {payload.name} added"}


@router.delete("/monitoring/alerts/rules/{name}", response_model=None)
def delete_alert_rule(name: str, _: dict = Depends(require_admin)) -> dict:
    try:
        alert_rules.delete_rule(name)
    except alert_rules.AlertRuleError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    return {"ok": True, "message": f"Alert rule {name} deleted"}


@router.get("/servers/{server_id}/monitoring", response_model=ServerMonitoringState)
async def server_monitoring_state(server_id: str, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ServerMonitoringState:
    server = _server_or_404(db, server_id, claims)
    scraped = False
    if server.metrics_enabled:
        try:
            result = await monitoring.prometheus_query(f'up{{server_id="{server.public_id}"}}')
            for series in (((result or {}).get("data") or {}).get("result") or []):
                value = series.get("value") or []
                if len(value) >= 2 and str(value[1]) == "1":
                    scraped = True
                    break
        except MonitoringError:
            # Prometheus not configured or unreachable -- report not-scraped rather than erroring
            scraped = False
    return ServerMonitoringState(
        metrics_enabled=bool(server.metrics_enabled),
        node_exporter_port=server.node_exporter_port,
        scraped=scraped,
        log_shipping_enabled=bool(server.log_shipping_enabled),
        log_sources=_stored_log_sources(server),
    )


def _stored_log_sources(server: Server) -> list[LogSourceEntry]:
    entries: list[LogSourceEntry] = []
    for item in _json_list(server.log_sources_json):
        name_or_path = str(item.get("name_or_path") or "").strip()
        if not name_or_path:
            continue
        source = str(item.get("source") or "journal").strip() or "journal"
        entries.append(LogSourceEntry(source=source, name_or_path=name_or_path))
    return entries


@router.post("/servers/{server_id}/vitals", response_model=ServerRead)
def refresh_vitals(server_id: str, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ServerRead:
    server = _server_or_404(db, server_id, claims)
    try:
        _apply_vitals(server, probe_vitals(server, _credentials_for(db, server,CredentialPayload())))
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    db.commit()
    db.refresh(server)
    return to_read(server)


# --- interactive shell -------------------------------------------------------------
# Highest-privilege surface in the product: a PTY as the stored SSH user, so it can do
# anything that user can. Admin-only (matching /services/restart), ACL-checked, capped in
# duration and idle time, and every session is written to audit_logs.
#
# All four limits below used to be module constants. They are settings now (SHELL_MAX_MINUTES,
# SHELL_IDLE_MINUTES, SHELL_MAX_SESSIONS, SHELL_MAX_SESSIONS_PER_USER) so an operator running a
# migration longer than the shipped cap can raise it without editing code. The values here are
# only the fallbacks used when the setting is missing or unparsable; config.py holds the real
# defaults and the reasoning behind the 480.
_SHELL_DEFAULT_MAX_MINUTES = 480
_SHELL_DEFAULT_IDLE_MINUTES = 30
# The shell is a tabbed workspace, so one operator legitimately holds several sessions. A single
# global cap therefore let one user starve everyone else: the ceiling is global *and* per-user,
# and the refusal says which one was hit because the two need different fixes.
_SHELL_DEFAULT_MAX_SESSIONS = 24
_SHELL_DEFAULT_MAX_PER_USER = 8
_shell_active = 0
_shell_active_by_user: dict[str, int] = {}
_shell_lock = threading.Lock()


def _shell_setting(name: str, fallback: int) -> int:
    # same defensive read as _max_war_bytes: a missing or unusable override must not take the
    # shell endpoint down, it falls back to the shipped default
    try:
        value = int(getattr(get_settings(), name, fallback))
    except (TypeError, ValueError):
        return fallback
    # a negative override is treated as 0, which on the time caps means "no limit" and on the
    # session counts means "none allowed" -- either way, never a negative threshold
    return max(0, value)


# None means the limit is disabled. This is the whole point of returning an Optional rather than
# a number: SHELL_MAX_MINUTES=0 / SHELL_IDLE_MINUTES=0 must mean "run indefinitely", and a 0 left
# in place as a threshold produces `now - last_activity > 0`, which is true on the very first
# tick and closes every session within seconds -- the exact opposite of what was asked for.
def _shell_limit_minutes(name: str, fallback_minutes: int) -> int | None:
    minutes = _shell_setting(name, fallback_minutes)
    return None if minutes <= 0 else minutes


def _shell_max_sessions() -> int:
    return _shell_setting("shell_max_sessions", _SHELL_DEFAULT_MAX_SESSIONS)


def _shell_max_per_user() -> int:
    return _shell_setting("shell_max_sessions_per_user", _SHELL_DEFAULT_MAX_PER_USER)


# Returns "" when a slot was taken, otherwise the close reason. Every caller that gets ""
# owes exactly one _shell_release_slot, from a finally block.
def _shell_claim_slot(actor: str) -> str:
    global _shell_active
    # read the settings outside the lock: nothing here mutates shared state
    max_per_user = _shell_max_per_user()
    max_sessions = _shell_max_sessions()
    with _shell_lock:
        # per-user first: when both caps are hit, closing one of your own tabs is the action
        # that actually helps, and "the server is full" would send the operator to the wrong place
        if _shell_active_by_user.get(actor, 0) >= max_per_user:
            if max_per_user == 0:
                return "Interactive shell is disabled (SHELL_MAX_SESSIONS_PER_USER=0)"
            return f"You already have {max_per_user} shell sessions open; close a tab first"
        if _shell_active >= max_sessions:
            if max_sessions == 0:
                return "Interactive shell is disabled (SHELL_MAX_SESSIONS=0)"
            return f"Server is at capacity ({max_sessions} shell sessions); try again shortly"
        _shell_active += 1
        _shell_active_by_user[actor] = _shell_active_by_user.get(actor, 0) + 1
    return ""


def _shell_release_slot(actor: str) -> None:
    global _shell_active
    with _shell_lock:
        _shell_active = max(0, _shell_active - 1)
        remaining = _shell_active_by_user.get(actor, 0) - 1
        # drop the key at zero rather than leaving a 0 entry per user who ever opened a shell
        if remaining > 0:
            _shell_active_by_user[actor] = remaining
        else:
            _shell_active_by_user.pop(actor, None)


def _audit(db: Session, actor: str, action: str, server: Server, details: str) -> None:
    db.add(
        AuditLog(
            actor=actor,
            action=action,
            resource_type="server",
            resource_id=server.public_id,
            details=details[:2000],
            server_id=server.id,
        )
    )
    db.commit()


@router.websocket("/servers/{server_id}/shell")
async def server_shell(websocket: WebSocket, server_id: str) -> None:
    global _shell_active
    await websocket.accept()

    # first frame must be the handshake: the socket is open but unauthenticated until then
    try:
        hello = await asyncio.wait_for(websocket.receive_json(), timeout=10)
    except Exception:
        await websocket.close(code=4401, reason="Handshake required")
        return

    token = str(hello.get("token") or "").strip()
    claims = decode_token(token) if token else None
    if not claims:
        # Desktop guest: an empty token from a loopback client in desktop mode opens a guest
        # session (same double gate as the HTTP path). A *present but invalid* token is still 4401.
        if not token and get_settings().desktop_mode and is_loopback_client(websocket.client.host if websocket.client else None):
            claims = guest_claims()
        else:
            await websocket.close(code=4401, reason="Invalid access token")
            return
    if claims.get("role") not in {"admin", "administrator"}:
        await websocket.close(code=4403, reason="Admin role required")
        return

    db = SessionLocal()
    session: ShellSession | None = None
    actor = str(claims.get("sub") or "unknown")
    started = datetime.now(timezone.utc)
    slot_held = False
    try:
        try:
            server = _server_or_404(db, server_id, claims)
            credentials = _credentials_for(db, server,CredentialPayload())
        except HTTPException as exc:
            await websocket.close(code=4404, reason=str(exc.detail)[:110])
            return

        refusal = _shell_claim_slot(actor)
        if refusal:
            await websocket.close(code=4429, reason=refusal[:110])
            return
        slot_held = True

        try:
            session = ShellSession(
                server,
                credentials,
                cols=_as_int(hello.get("cols"), 80),
                rows=_as_int(hello.get("rows"), 24),
            )
        except Exception as exc:
            # Exception, not just SshOperationError: paramiko raises its own connect and
            # authentication errors straight out of ShellSession, and letting one escape here
            # closed the socket with a bare 1011 and no reason -- the tab then showed a dead
            # terminal with nothing to act on. CancelledError is a BaseException, so a
            # shutdown mid-connect still propagates.
            await websocket.close(code=4500, reason=(str(exc) or type(exc).__name__)[:110])
            return

        _audit(db, actor, "shell.open", server, f"Interactive shell opened as {server.username}@{server.ip_address}")
        await websocket.send_text(f"\r\n\x1b[32m*** Connected to {server.hostname} as {server.username} ***\x1b[0m\r\n")

        sent = 0
        received = 0
        last_activity = asyncio.get_running_loop().time()
        loop = asyncio.get_running_loop()

        async def pump_remote_to_client() -> None:
            nonlocal sent, last_activity
            while True:
                data = await loop.run_in_executor(None, session.read)
                if data:
                    sent += len(data)
                    last_activity = loop.time()
                    await websocket.send_text(data.decode("utf-8", errors="replace"))
                else:
                    await asyncio.sleep(0.03)

        async def pump_client_to_remote() -> None:
            nonlocal received, last_activity
            while True:
                message = await websocket.receive_json()
                last_activity = loop.time()
                if "d" in message:
                    payload = str(message["d"])
                    received += len(payload)
                    session.write(payload)
                elif "r" in message and isinstance(message["r"], list) and len(message["r"]) == 2:
                    session.resize(_as_int(message["r"][0], 80), _as_int(message["r"][1], 24))

        async def enforce_limits() -> None:
            # Resolved once, when the session starts, so an operator editing .env cannot move
            # the goalposts under a shell that is already running.
            idle_minutes = _shell_limit_minutes("shell_idle_minutes", _SHELL_DEFAULT_IDLE_MINUTES)
            max_minutes = _shell_limit_minutes("shell_max_minutes", _SHELL_DEFAULT_MAX_MINUTES)
            if idle_minutes is None and max_minutes is None:
                # Both caps disabled: nothing to enforce, so park forever rather than return.
                # asyncio.wait(FIRST_EXCEPTION) below would not treat a clean return as a
                # reason to close, but parking keeps that independent of the wait mode.
                while True:
                    await asyncio.sleep(3600)
            # None -> no threshold at all. Never fall through to a 0 here: `now - last > 0` is
            # true on the first tick and would kill every session instantly.
            idle_seconds = None if idle_minutes is None else idle_minutes * 60
            max_seconds = None if max_minutes is None else max_minutes * 60
            while True:
                await asyncio.sleep(5)
                now = loop.time()
                # the reason names the setting that fired: an idle timeout and an absolute
                # timeout call for different responses from whoever is looking at the audit row
                if idle_seconds is not None and now - last_activity > idle_seconds:
                    raise TimeoutError(f"idle limit reached (SHELL_IDLE_MINUTES={idle_minutes})")
                if max_seconds is not None and (datetime.now(timezone.utc) - started).total_seconds() > max_seconds:
                    raise TimeoutError(f"absolute session limit reached (SHELL_MAX_MINUTES={max_minutes})")

        tasks = [asyncio.create_task(c()) for c in (pump_remote_to_client, pump_client_to_remote, enforce_limits)]
        reason = "closed"
        # The close record is written in `finally` so it survives every exit path, including
        # this handler being cancelled when the app shuts down mid-session. Writing it after
        # the try block instead would leave an unmatched shell.open row on that path, and
        # "every shell session is audited" would be false exactly when it matters.
        try:
            done, pending = await asyncio.wait(tasks, return_when=asyncio.FIRST_EXCEPTION)
            for task in pending:
                task.cancel()
            for task in done:
                if task.cancelled():
                    # .exception() re-raises CancelledError on a cancelled task
                    reason = "cancelled"
                    continue
                exc = task.exception()
                if isinstance(exc, TimeoutError):
                    reason = f"session ended: {exc}"
                elif isinstance(exc, (WebSocketDisconnect, SshOperationError)):
                    reason = "closed"
                elif exc is not None:
                    reason = f"error: {exc}"
        except asyncio.CancelledError:
            reason = "server shutting down"
            raise
        finally:
            for task in tasks:
                task.cancel()
            duration = int((datetime.now(timezone.utc) - started).total_seconds())
            try:
                _audit(
                    db,
                    actor,
                    "shell.close",
                    server,
                    f"Shell {reason} after {duration}s; {received} bytes from client, {sent} bytes from host",
                )
            except Exception:
                # never let an audit write failure mask the original reason for closing
                pass

        try:
            await websocket.close()
        except Exception:
            pass
    finally:
        if session is not None:
            session.close()
        # released on the slot flag, not on `session`: the SSH connection can fail *after* the
        # slot was claimed, and keying this off the session object leaked a slot per failed
        # connect until the process restarted
        if slot_held:
            _shell_release_slot(actor)
        db.close()


# --- SFTP file transfer ------------------------------------------------------------
# Same privilege as the shell above: arbitrary filesystem access as the stored SSH user, so
# require_admin plus the per-server ACL, exactly like /shell.
#
# Paths are deliberately NOT jailed to a subtree. The same credential already gets a full PTY
# on the same host through the shell endpoint, so a jail here would block nothing an operator
# cannot already do and would only mislead the next reader into thinking this surface is
# narrower than it is. What is enforced instead, and is worth enforcing: malformed input is
# rejected, `.`/`..` are collapsed so the path shown, audited and sent are one string, sizes
# are capped in both directions from the remote stat, and every byte moved is audited.
_SFTP_PATH_MAX = 4096
_SFTP_NAME_MAX = 255
_SFTP_STREAM_CHUNK = 256 * 1024
_DEFAULT_MAX_DOWNLOAD_MB = 200


def _max_download_bytes() -> int:
    # mirrors _max_war_bytes: a missing or unusable MAX_DOWNLOAD_MB must still leave a cap in
    # force rather than allow an unbounded read through the API process
    try:
        limit_mb = int(getattr(get_settings(), "max_download_mb", _DEFAULT_MAX_DOWNLOAD_MB))
    except (TypeError, ValueError):
        limit_mb = _DEFAULT_MAX_DOWNLOAD_MB
    return max(1, limit_mb) * 1024 * 1024


def _sftp_too_large(size_bytes: int, limit_bytes: int, setting: str) -> str:
    return f"{size_bytes} bytes exceeds the {limit_bytes // (1024 * 1024)} MB limit ({setting})"


def _sftp_path(raw: str, *, required: bool) -> str:
    path = (raw or "").strip()
    if "\x00" in path:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Remote path must not contain NUL bytes")
    if len(path) > _SFTP_PATH_MAX:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Remote path must be at most {_SFTP_PATH_MAX} characters")
    if not path:
        if required:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A remote path is required")
        # empty means "the SSH user's home", resolved from the SFTP session itself
        return ""
    # posixpath, not os.path: a backslash is a legal character in a Linux file name and must
    # survive normalisation untouched, which os.path on Windows would not do
    normalised = posixpath.normpath(path)
    if normalised.startswith("//"):
        normalised = "/" + normalised.lstrip("/")
    return normalised


def _sftp_upload_name(candidate: str) -> str:
    # The target directory is the operator's to choose; the name must not add path structure of
    # its own. Rejected rather than reduced to its basename: silently turning `../x` into `x`
    # writes a file the client did not ask for, and the same discipline as _validated_war_name
    # keeps the created path exactly what the request named.
    name = (candidate or "").strip()
    if not name or name in {".", ".."}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A file name is required")
    if len(name) > _SFTP_NAME_MAX:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"File name must be at most {_SFTP_NAME_MAX} characters")
    if "/" in name or "\\" in name or ".." in name or any(ord(char) < 32 or ord(char) == 127 for char in name):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File name must be a bare file name with no directory separators, '..' or control characters",
        )
    return name


def _sftp_download_headers(remote_path: str) -> dict[str, str]:
    base = posixpath.basename(remote_path.rstrip("/")) or "download"
    # A remote file name may legally contain newlines, quotes and backslashes, any of which
    # would break out of the header value. The ASCII fallback is sanitised character by
    # character; filename* is percent-encoded, which makes CR/LF unrepresentable.
    ascii_name = "".join(char if 32 <= ord(char) < 127 and char not in '"\\' else "_" for char in base) or "download"
    return {"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(base, safe='')}"}


def _sftp_entries(entries: object) -> list[dict]:
    items = [entry for entry in entries if isinstance(entry, dict)] if isinstance(entries, list) else []
    return sorted(items, key=lambda entry: (str(entry.get("type") or "") != "dir", str(entry.get("name") or "").lower()))


def _sftp_body(data: bytes) -> Iterator[bytes]:
    # the ssh_ops primitive returns bytes, so the file is already resident; chunking here keeps
    # Starlette from holding a second full-size copy while it writes the response
    view = memoryview(data)
    for start in range(0, len(view), _SFTP_STREAM_CHUNK):
        yield bytes(view[start : start + _SFTP_STREAM_CHUNK])


@router.post("/servers/{server_id}/sftp/list", response_model=SftpListing)
def sftp_list_api(
    server_id: str,
    path: str = Body("", embed=True),
    claims: dict = Depends(require_admin),
    db: Session = Depends(get_db),
) -> SftpListing:
    server = _server_or_404(db, server_id, claims)
    target = _sftp_path(path, required=False)
    try:
        listing = sftp_list(server, _credentials_for(db, server,CredentialPayload()), target)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    # deliberately not audited: a listing moves no data off the host, and a row per navigation
    # step would bury the transfers that do matter
    return SftpListing.model_validate({**listing, "entries": _sftp_entries(listing.get("entries"))})


@router.post("/servers/{server_id}/sftp/download")
def sftp_download_api(
    server_id: str,
    path: str = Body(..., embed=True),
    claims: dict = Depends(require_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    server = _server_or_404(db, server_id, claims)
    target = _sftp_path(path, required=True)
    limit = _max_download_bytes()
    credentials = _credentials_for(db, server,CredentialPayload())
    try:
        info = sftp_stat(server, credentials, target)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    # a symlink reports type "link" with the resolved kind in target_type, and downloading a
    # link means downloading what it points at, so both have to be checked
    if "dir" in {str(info.get("type") or ""), str(info.get("target_type") or "")}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="That path is a directory, not a file")
    # from the stat, before any bytes are read: learning the size after pulling 4 GB through
    # this process is exactly the failure the cap exists to prevent
    size = _as_int(info.get("size_bytes"), 0)
    if size > limit:
        raise HTTPException(status_code=_HTTP_413, detail=_sftp_too_large(size, limit, "MAX_DOWNLOAD_MB"))
    try:
        data = sftp_download(server, credentials, target, limit)
    except SftpTooLarge as exc:
        # the file grew between the stat above and the read, or reported st_size 0 as /proc
        # does; still a 413, not a generic failure
        raise HTTPException(status_code=_HTTP_413, detail=str(exc)) from exc
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    _audit(
        db,
        str(claims.get("sub") or "unknown"),
        "sftp.download",
        server,
        f"SFTP download {target} ({len(data)} bytes) as {server.username}@{server.ip_address}",
    )
    headers = _sftp_download_headers(target)
    headers["Content-Length"] = str(len(data))
    return StreamingResponse(_sftp_body(data), media_type="application/octet-stream", headers=headers)


@router.post("/servers/{server_id}/sftp/upload", response_model=SftpUploadResult)
def sftp_upload_api(
    server_id: str,
    path: str = Form(..., max_length=_SFTP_PATH_MAX + 1),
    file: UploadFile = File(...),
    claims: dict = Depends(require_admin),
    db: Session = Depends(get_db),
) -> SftpUploadResult:
    server = _server_or_404(db, server_id, claims)
    # local checks first: never open an SSH connection for a request that cannot be valid
    target_dir = _sftp_path(path, required=True)
    name = _sftp_upload_name(file.filename or "")
    limit = _max_war_bytes()
    try:
        # the message cannot quote the real size: when the client sends no Content-Length for
        # the part, the cap is hit mid-read and the total is unknown at that point
        data = _read_capped_upload(file, limit, f"Upload exceeds the {limit // (1024 * 1024)} MB limit (MAX_WAR_MB)")
    finally:
        file.file.close()
    if not data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")
    try:
        result = sftp_upload(server, _credentials_for(db, server,CredentialPayload()), target_dir, name, data)
    except SshOperationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    written = _deploy_int(result, "bytes_written", len(data))
    remote_path = _deploy_text(result, "path", "target_path") or f"{target_dir.rstrip('/')}/{name}"
    _audit(
        db,
        str(claims.get("sub") or "unknown"),
        "sftp.upload",
        server,
        f"SFTP upload {remote_path} ({written} bytes) as {server.username}@{server.ip_address}",
    )
    return SftpUploadResult(ok=True, message=f"Uploaded {name} to {target_dir}", path=remote_path, bytes_written=written)


_SFTP_DELETE_LABELS = {"dir": "directory", "link": "symlink", "file": "file"}


@router.post("/servers/{server_id}/sftp/delete", response_model=SftpDeleteResult)
def sftp_delete_api(
    server_id: str,
    payload: SftpDeleteRequest,
    claims: dict = Depends(require_admin),
    db: Session = Depends(get_db),
) -> SftpDeleteResult:
    # The first operation in the product that destroys data on a managed host. Same gate as the
    # rest of this section (require_admin + the per-server ACL), and unlike /sftp/list it is
    # always audited: the unaudited restarts are already a documented gap, and an unaudited
    # delete would be strictly worse -- a removed file leaves nothing behind to reconstruct
    # who removed it or when.
    server = _server_or_404(db, server_id, claims)
    target = _sftp_path(payload.path, required=True)
    try:
        result = sftp_delete(server, _credentials_for(db, server,CredentialPayload()), target, recursive=payload.recursive)
    except SshOperationError as exc:
        # 400 with the message intact, never a rewritten one. Everything ssh_ops refuses arrives
        # this way -- a directory without recursive=True, a shallow recursive target like /etc,
        # `/` itself, permission denied, not found, directory not empty -- and each of those
        # needs a different action from the operator. Flattening them into a generic failure
        # would leave the UI with nothing useful to say.
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    removed_path = _deploy_text(result, "path") or target
    kind = _deploy_text(result, "deleted") or "file"
    entries = _deploy_int(result, "entries_removed", 1)
    _audit(
        db,
        str(claims.get("sub") or "unknown"),
        "sftp.delete",
        server,
        f"SFTP delete {removed_path} ({kind}, recursive={payload.recursive}, entries_removed={entries}) "
        f"as {server.username}@{server.ip_address}",
    )
    label = _SFTP_DELETE_LABELS.get(kind, kind or "file")
    message = f"Deleted {label} {removed_path}"
    if kind == "dir":
        message = f"{message} ({entries} entries removed)"
    return SftpDeleteResult(
        ok=True,
        message=message,
        path=removed_path,
        deleted=kind,
        recursive=payload.recursive,
        entries_removed=entries,
    )


# --- shell favorites ---------------------------------------------------------------
# Personal saved commands, so require_user rather than require_admin: a stored string does
# nothing on its own and can only run in a shell the user already has.
#
# Every query below filters on the calling user's id in the WHERE clause rather than loading
# by primary key and comparing afterwards, so another user's id is simply not found. The 404
# on someone else's row is deliberate: a 403 would confirm that the id exists.
def _favorite_read(favorite: ShellFavorite) -> ShellFavoriteRead:
    return ShellFavoriteRead.model_validate(favorite, from_attributes=True)


def _favorite_conflict(name: str) -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"You already have a favorite named {name}")


@router.get("/shell/favorites", response_model=list[ShellFavoriteRead])
def list_shell_favorites(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[ShellFavoriteRead]:
    user = _current_user(db, claims)
    favorites = db.scalars(
        select(ShellFavorite)
        .where(ShellFavorite.user_id == user.id)
        # id breaks the tie: created_at has one-second resolution on SQLite, so several
        # favorites saved in the same second would otherwise come back in arbitrary order
        .order_by(ShellFavorite.created_at.desc(), ShellFavorite.id.desc())
    ).all()
    return [_favorite_read(favorite) for favorite in favorites]


@router.post("/shell/favorites", response_model=ShellFavoriteRead, status_code=status.HTTP_201_CREATED)
def create_shell_favorite(payload: ShellFavoriteCreate, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> ShellFavoriteRead:
    user = _current_user(db, claims)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A favorite name is required")
    # the command is stored exactly as typed. There is nothing to sanitise: it is text destined
    # for the user's own shell, and a validator here would only reject legitimate one-liners.
    if not payload.command.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A command is required")
    if db.scalar(select(ShellFavorite).where(ShellFavorite.user_id == user.id, ShellFavorite.name == name)):
        raise _favorite_conflict(name)
    favorite = ShellFavorite(user_id=user.id, name=name, command=payload.command)
    db.add(favorite)
    try:
        db.commit()
    except IntegrityError as exc:
        # the check above loses to a concurrent insert; uq_shell_favorites_user_name is what
        # actually guarantees uniqueness, so report the conflict instead of a 500
        db.rollback()
        raise _favorite_conflict(name) from exc
    db.refresh(favorite)
    return _favorite_read(favorite)


@router.delete("/shell/favorites/{favorite_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_shell_favorite(favorite_id: int, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> None:
    user = _current_user(db, claims)
    favorite = db.scalar(select(ShellFavorite).where(ShellFavorite.id == favorite_id, ShellFavorite.user_id == user.id))
    if not favorite:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Favorite not found")
    db.delete(favorite)
    db.commit()


# --- database console (feature/db-connect) -------------------------------------------------
#
# A standalone SQL console: a signed-in (non-guest) user supplies connection parameters and
# credentials per request, and the backend opens a fresh PostgreSQL/MySQL connection, runs the
# work, and closes it. Nothing is stored. Guarded by require_user, which by design (2026-08-14)
# admits the desktop guest -- the guest role is admin and its menu includes the DB console. Only
# a *_not_guest guard would exclude the guest, and these routes intentionally don't use one.
#
# IN scope: standalone params (not a managed server), test + run a read-only query, postgres +
# mysql, per-request credentials. OUT of scope (follow-ons): saved/stored connections, a schema
# browser, SQL autocomplete, and write-transaction management.

# how many rows a single query may return by default. A top-level LIMIT in the SQL raises the
# cap to the user's own LIMIT instead (hard-capped at _DB_MAX_ROW_LIMIT); see parse_limit.
_DB_DEFAULT_ROW_LIMIT = 200
_DB_MAX_ROW_LIMIT = 5000


def _db_engine_or_400(engine: str) -> str:
    normalized = (engine or "").strip().lower()
    if normalized not in ENGINES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Engine must be postgres, mysql, sqlite, or mssql")
    return normalized


@router.post("/db/test-connection", response_model=DbConnectionResult)
def db_test_connection(payload: DbConnectionRequest, _: dict = Depends(require_user)) -> DbConnectionResult:
    engine = _db_engine_or_400(payload.engine)
    port = payload.port or DEFAULT_PORTS.get(engine, 0)
    try:
        message = db_console.test_connection(engine, payload.host, port, payload.username, payload.password, payload.database)
        return DbConnectionResult(ok=True, message=message)
    except DbConsoleError as exc:
        # a bad host/credentials is an expected outcome here, not a server error: report it as a
        # clean ok=false with HTTP 200 so the UI can render the reason inline
        return DbConnectionResult(ok=False, message=str(exc))


@router.post("/db/query", response_model=DbQueryResult)
def db_query(payload: DbQueryRequest, _: dict = Depends(require_user)) -> DbQueryResult:
    engine = _db_engine_or_400(payload.engine)
    port = payload.port or DEFAULT_PORTS.get(engine, 0)
    row_cap = min(db_console.parse_limit(payload.sql) or _DB_DEFAULT_ROW_LIMIT, _DB_MAX_ROW_LIMIT)
    try:
        result = db_console.run_query(engine, payload.host, port, payload.username, payload.password, payload.database, payload.sql, row_cap)
    except DbConsoleError as exc:
        # never a 500 on a bad query or unreachable host: surface the DB error text as a 400 the
        # console can display
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return DbQueryResult(**result)


# --- saved database connections (feature/db-connect follow-on) ------------------------------
#
# Persistent, reusable versions of the ad-hoc console above: the connection parameters are stored,
# the password is encrypted at rest with the same Fernet path (never echoed -- has_password says
# whether one is stored), and `group` is a folder NAME resolved to a Folder exactly as servers and
# clusters are grouped. The API speaks in the public_id, never the autoincrement key. All endpoints
# require a real signed-in (non-guest) user, matching the ad-hoc /db routes.


def _db_connection_or_404(db: Session, connection_id: str) -> DbConnection:
    conn = db.scalar(select(DbConnection).where(DbConnection.public_id == connection_id))
    if not conn and connection_id.isdigit():
        conn = db.get(DbConnection, int(connection_id))
    if not conn:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Connection not found")
    return conn


def _db_connection_group_name(db: Session, conn: DbConnection) -> str | None:
    if conn.folder_id is None:
        return None
    folder = db.get(Folder, conn.folder_id)
    return folder.name if folder else None


def _db_connection_read(db: Session, conn: DbConnection) -> DbConnectionRead:
    return DbConnectionRead(
        id=conn.public_id,
        name=conn.name,
        engine=conn.engine,
        host=conn.host,
        port=conn.port,
        username=conn.username,
        database=conn.database,
        environment=conn.environment or "",
        show_all_databases=bool(conn.show_all_databases),
        group=_db_connection_group_name(db, conn),
        has_password=bool(conn.encrypted_password),
        created_at=conn.created_at,
    )


@router.get("/db/connections", response_model=list[DbConnectionRead])
def list_db_connections(_: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbConnectionRead]:
    conns = db.scalars(select(DbConnection).order_by(func.lower(DbConnection.name))).all()
    return [_db_connection_read(db, conn) for conn in conns]


def _validate_db_connection_target(engine: str, host: str, database: str) -> None:
    """Engine-specific required-field check: sqlite needs a database file path (host is ignored);
    the networked engines need a host. Raises HTTP 400 on a missing value."""
    if engine == "sqlite":
        if not (database or "").strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SQLite requires a database file path")
    elif not (host or "").strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Host is required")


@router.post("/db/connections", response_model=DbConnectionRead, status_code=status.HTTP_201_CREATED)
def create_db_connection(payload: DbConnectionCreate, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbConnectionRead:
    engine = _db_engine_or_400(payload.engine)
    _validate_db_connection_target(engine, payload.host, payload.database)
    conn = DbConnection(
        public_id=str(uuid.uuid4()),
        name=payload.name.strip(),
        engine=engine,
        host=payload.host.strip(),
        port=payload.port or DEFAULT_PORTS.get(engine, 0),
        username=payload.username,
        encrypted_password=encrypt_secret(payload.password),
        database=payload.database,
        environment=(payload.environment or "").strip(),
        show_all_databases=bool(payload.show_all_databases),
        folder_id=_resolve_group_to_folder_id(db, payload.group),
    )
    db.add(conn)
    db.commit()
    db.refresh(conn)
    return _db_connection_read(db, conn)


@router.post("/db/connections/test", response_model=DbConnectionResult)
def test_unsaved_db_connection(payload: DbConnectionCreate, _: dict = Depends(require_user)) -> DbConnectionResult:
    # Test the parameters as typed, without persisting anything. A bad host/credentials is an
    # expected outcome, reported as ok=false with HTTP 200 (like the ad-hoc db_test_connection).
    engine = _db_engine_or_400(payload.engine)
    port = payload.port or DEFAULT_PORTS.get(engine, 0)
    try:
        message = db_console.test_connection(engine, payload.host, port, payload.username, payload.password, payload.database)
        return DbConnectionResult(ok=True, message=message)
    except DbConsoleError as exc:
        return DbConnectionResult(ok=False, message=str(exc))


@router.get("/db/connections/{connection_id}", response_model=DbConnectionRead)
def get_db_connection(connection_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbConnectionRead:
    return _db_connection_read(db, _db_connection_or_404(db, connection_id))


@router.patch("/db/connections/{connection_id}", response_model=DbConnectionRead)
def update_db_connection(connection_id: str, payload: DbConnectionUpdate, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbConnectionRead:
    conn = _db_connection_or_404(db, connection_id)
    data = payload.model_dump(exclude_unset=True)

    if "engine" in data and data["engine"] is not None:
        conn.engine = _db_engine_or_400(data["engine"])
    if "name" in data and data["name"] is not None:
        name = data["name"].strip()
        if not name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Connection name cannot be empty")
        conn.name = name
    if "host" in data and data["host"] is not None:
        host = data["host"].strip()
        # sqlite has no host (its `database` is a local file path), so an empty host is valid there;
        # the networked engines still require one.
        if not host and conn.engine != "sqlite":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Host cannot be empty")
        conn.host = host
    if "port" in data and data["port"] is not None:
        conn.port = data["port"]
    if "username" in data and data["username"] is not None:
        conn.username = data["username"]
    if "database" in data and data["database"] is not None:
        conn.database = data["database"]
    if "environment" in data and data["environment"] is not None:
        conn.environment = data["environment"].strip()
    if "show_all_databases" in data and data["show_all_databases"] is not None:
        conn.show_all_databases = bool(data["show_all_databases"])
    # password: only overwrite when a non-empty value is supplied, so an edit that leaves the box
    # blank keeps the stored credential instead of wiping it (mirrors the kube secret handling).
    if data.get("password"):
        conn.encrypted_password = encrypt_secret(data["password"])
    if "group" in data:
        conn.folder_id = _resolve_group_to_folder_id(db, data["group"])

    db.commit()
    db.refresh(conn)
    return _db_connection_read(db, conn)


@router.delete("/db/connections/{connection_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_db_connection(connection_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> None:
    conn = _db_connection_or_404(db, connection_id)
    db.delete(conn)
    db.commit()


@router.post("/db/connections/{connection_id}/test", response_model=DbConnectionResult)
def test_saved_db_connection(connection_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbConnectionResult:
    conn = _db_connection_or_404(db, connection_id)
    try:
        message = db_console.test_connection(
            conn.engine, conn.host, conn.port, conn.username, decrypt_secret(conn.encrypted_password), conn.database
        )
        return DbConnectionResult(ok=True, message=message)
    except DbConsoleError as exc:
        return DbConnectionResult(ok=False, message=str(exc))


@router.get("/db/connections/{connection_id}/tables", response_model=list[DbTable])
def list_db_connection_tables(connection_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbTable]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        tables = db_console.list_tables(
            conn.engine, conn.host, conn.port, conn.username, decrypt_secret(conn.encrypted_password), conn.database
        )
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbTable(**table) for table in tables]


def _record_db_query_history(db: Session, conn: DbConnection, user_email: str, sql: str, *, status_: str, error: str, row_count: int, elapsed_ms: int) -> None:
    """Insert one DbQueryHistory row, snapshotting the connection's identity.

    Best-effort: a failure to record history must never turn a successful (or already-failed)
    query into an HTTP 500, so any exception is swallowed after rolling the session back.
    """
    try:
        db.add(DbQueryHistory(
            connection_id=conn.id,
            connection_name=conn.name,
            engine=conn.engine,
            database=conn.database or "",
            user_email=user_email or "",
            sql=sql,
            status=status_,
            error=error,
            row_count=row_count,
            elapsed_ms=elapsed_ms,
        ))
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass


@router.post("/db/connections/{connection_id}/query", response_model=DbQueryResult)
def query_db_connection(connection_id: str, payload: DbConnectionQueryRequest, database: str | None = None, claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbQueryResult:
    conn = _db_connection_or_404(db, connection_id)
    user_email = claims.get("sub", "")
    # Optional ?database= override lets the caller run the query against a different database on the
    # same server without editing the saved connection; blank/None keeps the stored database.
    effective_database = database or conn.database
    row_cap = min(db_console.parse_limit(payload.sql) or _DB_DEFAULT_ROW_LIMIT, _DB_MAX_ROW_LIMIT)
    try:
        result = db_console.run_query(
            conn.engine, conn.host, conn.port, conn.username, decrypt_secret(conn.encrypted_password), effective_database, payload.sql, row_cap
        )
    except DbConsoleError as exc:
        _record_db_query_history(
            db, conn, user_email, payload.sql,
            status_="error", error=str(exc), row_count=0, elapsed_ms=0,
        )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    _record_db_query_history(
        db, conn, user_email, payload.sql,
        status_="success", error="", row_count=int(result.get("row_count", 0)), elapsed_ms=int(result.get("elapsed_ms", 0)),
    )
    return DbQueryResult(**result)


# --- database metadata / catalog browsing (DBeaver-style) -----------------------------------
#
# Read-only introspection over a saved connection: the schema tree (schemas -> tables/views/
# routines) and the per-table detail panes (columns/indexes/constraints/foreign-keys), plus a
# generate-SQL helper that builds boilerplate from a table's real columns. Every call decrypts the
# stored password, runs one short introspection query via db_metadata, and surfaces a DbConsoleError
# as a clean HTTP 400. Same auth as the rest of the /db routes: require_user, which by design
# admits the desktop guest (no *_not_guest guard is applied).


def _db_meta_args(conn: DbConnection, database: str | None = None) -> tuple:
    """The leading (engine, host, port, username, password, database) argument tuple every
    db_metadata function takes, with the stored password decrypted.

    `database` is an optional override so the caller can browse a DIFFERENT database on the same
    server (postgres has separate schemas per database; mysql treats database == schema; sqlite
    ignores it) without editing the saved connection. Blank/None falls back to the stored database.
    """
    effective_database = database or conn.database
    return (conn.engine, conn.host, conn.port, conn.username, decrypt_secret(conn.encrypted_password), effective_database)


@router.get("/db/connections/{connection_id}/databases", response_model=list[DbDatabase])
def list_db_connection_databases(connection_id: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbDatabase]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_databases(*_db_meta_args(conn, database))
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbDatabase(**row) for row in rows]


@router.get("/db/connections/{connection_id}/schemas", response_model=list[DbSchema])
def list_db_connection_schemas(connection_id: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbSchema]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_schemas(*_db_meta_args(conn, database))
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbSchema(**row) for row in rows]


@router.get("/db/connections/{connection_id}/schemas/{schema}/tables", response_model=list[DbTable])
def list_db_connection_schema_tables(connection_id: str, schema: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbTable]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_tables_in_schema(*_db_meta_args(conn, database), schema)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbTable(**row) for row in rows]


@router.get("/db/connections/{connection_id}/schemas/{schema}/routines", response_model=list[DbRoutine])
def list_db_connection_schema_routines(connection_id: str, schema: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbRoutine]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_routines(*_db_meta_args(conn, database), schema)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbRoutine(**row) for row in rows]


@router.get("/db/connections/{connection_id}/tables/{schema}/{table}/columns", response_model=list[DbColumn])
def list_db_connection_table_columns(connection_id: str, schema: str, table: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbColumn]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_columns(*_db_meta_args(conn, database), schema, table)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbColumn(**row) for row in rows]


@router.get("/db/connections/{connection_id}/tables/{schema}/{table}/indexes", response_model=list[DbIndex])
def list_db_connection_table_indexes(connection_id: str, schema: str, table: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbIndex]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_indexes(*_db_meta_args(conn, database), schema, table)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbIndex(**row) for row in rows]


@router.get("/db/connections/{connection_id}/tables/{schema}/{table}/constraints", response_model=list[DbConstraint])
def list_db_connection_table_constraints(connection_id: str, schema: str, table: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbConstraint]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_constraints(*_db_meta_args(conn, database), schema, table)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbConstraint(**row) for row in rows]


@router.get("/db/connections/{connection_id}/tables/{schema}/{table}/foreign-keys", response_model=list[DbForeignKey])
def list_db_connection_table_foreign_keys(connection_id: str, schema: str, table: str, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[DbForeignKey]:
    conn = _db_connection_or_404(db, connection_id)
    try:
        rows = db_metadata.list_foreign_keys(*_db_meta_args(conn, database), schema, table)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [DbForeignKey(**row) for row in rows]


@router.post("/db/connections/{connection_id}/generate-sql", response_model=DbGenerateSqlResult)
def generate_db_connection_sql(connection_id: str, payload: DbGenerateSqlRequest, database: str | None = None, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> DbGenerateSqlResult:
    conn = _db_connection_or_404(db, connection_id)
    try:
        sql = db_metadata.generate_sql(*_db_meta_args(conn, database), payload.schema, payload.table, payload.kind)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return DbGenerateSqlResult(sql=sql)


# --- schema rename / backup / restore (admin DDL + pg client) -------------------------------
#
# Mutating operations, so unlike the read-only browsing routes above these require an admin (not a
# desktop guest). Schema rename is PostgreSQL-only DDL; backup/restore shell out to the pg client
# binaries via app.services.db_backup. Every other engine returns a clean 400.

# A schema name is a plain SQL identifier once validated; this is stricter than the SQL grammar on
# purpose (no spaces, no quotes) so the value can be safely quoted into an ALTER SCHEMA statement.
_DB_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
# hard cap on an uploaded restore file, enforced while streaming it to a temp file.
_DB_RESTORE_MAX_BYTES = 200 * 1024 * 1024


@router.post("/db/connections/{connection_id}/schemas/{schema}/rename", response_model=ActionResult)
def rename_db_connection_schema(
    connection_id: str,
    schema: str,
    payload: DbSchemaRenameRequest,
    database: str | None = None,
    claims: dict = Depends(require_admin_not_guest),
    db: Session = Depends(get_db),
) -> ActionResult:
    conn = _db_connection_or_404(db, connection_id)
    if conn.engine != "postgres":
        # mysql schemas are databases; sqlite/mssql schema rename is not offered here.
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Renaming a schema is not supported for {conn.engine}")
    new_name = payload.new_name.strip()
    if not _DB_IDENT_RE.match(new_name) or len(new_name) > 63:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="New schema name must be a plain identifier (letters, digits, underscore)")
    # new_name is validated above; schema comes from the URL. Both are double-quoted (with any
    # embedded quote doubled) so the statement is a safe, exact identifier rename.
    quoted_old = '"' + schema.replace('"', '""') + '"'
    quoted_new = '"' + new_name.replace('"', '""') + '"'
    sql = f"ALTER SCHEMA {quoted_old} RENAME TO {quoted_new}"
    try:
        db_console.execute_ddl(*_db_meta_args(conn, database), sql)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    db.add(AuditLog(
        actor=claims.get("sub", ""),
        action="db.schema.rename",
        resource_type="db_connection",
        resource_id=conn.public_id,
        details=f"Renamed schema {schema} to {new_name} on {conn.name}"[:2000],
        server_id=None,
    ))
    db.commit()
    return ActionResult(ok=True, message=f"Schema {schema} renamed to {new_name}")


def _db_backup_fields(conn: DbConnection) -> dict:
    return {
        "host": conn.host,
        "port": conn.port,
        "username": conn.username,
        "password": decrypt_secret(conn.encrypted_password),
    }


@router.post("/db/connections/{connection_id}/backup")
def backup_db_connection(
    connection_id: str,
    database: str | None = None,
    claims: dict = Depends(require_admin_not_guest),
    db: Session = Depends(get_db),
) -> FileResponse:
    conn = _db_connection_or_404(db, connection_id)
    if conn.engine != "postgres":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Backup is not supported for {conn.engine}")
    effective_database = database or conn.database
    try:
        path, filename = db_backup.backup(_db_backup_fields(conn), effective_database)
    except DbConsoleError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    db.add(AuditLog(
        actor=claims.get("sub", ""),
        action="db.backup",
        resource_type="db_connection",
        resource_id=conn.public_id,
        details=f"Backed up database {effective_database} on {conn.name}"[:2000],
        server_id=None,
    ))
    db.commit()
    # stream the dump as an attachment and delete the temp file once the response is finished.
    return FileResponse(
        path,
        media_type="application/octet-stream",
        filename=filename,
        background=BackgroundTask(db_backup._safe_remove, path),
    )


@router.post("/db/connections/{connection_id}/restore", response_model=ActionResult)
def restore_db_connection(
    connection_id: str,
    database: str | None = None,
    file: UploadFile = File(...),
    claims: dict = Depends(require_admin_not_guest),
    db: Session = Depends(get_db),
) -> ActionResult:
    conn = _db_connection_or_404(db, connection_id)
    if conn.engine != "postgres":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Restore is not supported for {conn.engine}")
    effective_database = database or conn.database

    # stream the upload to a temp file with a hard size cap, so a huge upload cannot exhaust memory
    # or disk. The temp file is always removed, success or failure.
    fd, tmp_path = tempfile.mkstemp(prefix="pgrestore_", suffix=".dump")
    total = 0
    try:
        with os.fdopen(fd, "wb") as out:
            while True:
                chunk = file.file.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > _DB_RESTORE_MAX_BYTES:
                    raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Upload exceeds the maximum restore size")
                out.write(chunk)
        if total == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")
        try:
            message = db_backup.restore(_db_backup_fields(conn), effective_database, tmp_path)
        except DbConsoleError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    finally:
        db_backup._safe_remove(tmp_path)

    db.add(AuditLog(
        actor=claims.get("sub", ""),
        action="db.restore",
        resource_type="db_connection",
        resource_id=conn.public_id,
        details=f"Restored database {effective_database} on {conn.name} ({total} bytes)"[:2000],
        server_id=None,
    ))
    db.commit()
    return ActionResult(ok=True, message=message)


# --- database query history (feature/db-connect follow-on) ----------------------------------
#
# A per-user log of queries run through the saved-connection console (recorded by the query route
# above). Reads are scoped to the caller; DELETE clears only the caller's own rows. Nothing secret
# is stored -- only the SQL text and its outcome.

_DB_HISTORY_DEFAULT_LIMIT = 100
_DB_HISTORY_MAX_LIMIT = 500


def _db_query_history_read(row: DbQueryHistory) -> DbQueryHistoryRead:
    return DbQueryHistoryRead(
        id=row.id,
        connection_id=row.connection_id,
        connection_name=row.connection_name,
        engine=row.engine,
        database=row.database,
        user_email=row.user_email,
        sql=row.sql,
        status=row.status,
        error=row.error,
        row_count=row.row_count,
        elapsed_ms=row.elapsed_ms,
        created_at=row.created_at,
    )


@router.get("/db/query-history", response_model=list[DbQueryHistoryRead])
def list_db_query_history(
    connection_id: int | None = None,
    status: str | None = None,
    search: str | None = None,
    limit: int = _DB_HISTORY_DEFAULT_LIMIT,
    claims: dict = Depends(require_user),
    db: Session = Depends(get_db),
) -> list[DbQueryHistoryRead]:
    # scoped to the caller: history is a personal record, not a shared audit log
    stmt = select(DbQueryHistory).where(DbQueryHistory.user_email == claims.get("sub", ""))
    if connection_id is not None:
        stmt = stmt.where(DbQueryHistory.connection_id == connection_id)
    if status:
        stmt = stmt.where(DbQueryHistory.status == status.strip().lower())
    if search:
        stmt = stmt.where(func.lower(DbQueryHistory.sql).like(f"%{search.strip().lower()}%"))
    capped = max(1, min(limit or _DB_HISTORY_DEFAULT_LIMIT, _DB_HISTORY_MAX_LIMIT))
    stmt = stmt.order_by(DbQueryHistory.id.desc()).limit(capped)
    return [_db_query_history_read(row) for row in db.scalars(stmt).all()]


@router.delete("/db/query-history", status_code=status.HTTP_204_NO_CONTENT)
def clear_db_query_history(claims: dict = Depends(require_user), db: Session = Depends(get_db)) -> None:
    # clear only the caller's own history
    db.execute(delete(DbQueryHistory).where(DbQueryHistory.user_email == claims.get("sub", "")))
    db.commit()


# --- kubernetes clusters (feature/kubernetes) ----------------------------------------------
#
# A saved Kubernetes cluster is persisted much like a Server: the API speaks in its public_id, its
# credentials (a full kubeconfig YAML, or a bearer token) are encrypted at rest with the same
# Fernet path, and the CA certificate is a public cert stored as plain text. `group` is a folder
# NAME, resolved to a Folder exactly as servers are grouped. Writes are admin-only; the live reads
# (nodes/pods/...) use require_user, which by design admits the desktop guest (mirroring the DB
# console -- no *_not_guest guard); the mutating actions (restart/scale/cordon/delete) require a
# non-guest admin.

_KUBE_AUTH_METHODS = {"kubeconfig", "token"}


def _kube_auth_or_400(value: str) -> str:
    method = (value or "").strip().lower()
    if method not in _KUBE_AUTH_METHODS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="auth_method must be 'kubeconfig' or 'token'")
    return method


def _cluster_or_404(db: Session, cluster_id: str) -> KubeCluster:
    cluster = db.scalar(select(KubeCluster).where(KubeCluster.public_id == cluster_id))
    if not cluster and cluster_id.isdigit():
        cluster = db.get(KubeCluster, int(cluster_id))
    if not cluster:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cluster not found")
    return cluster


def _resolve_group_to_folder_id(db: Session, group: str | None) -> int | None:
    # group is a folder NAME. Match an existing folder case-insensitively; create one if absent so
    # a cluster can be grouped in a single call, mirroring how folders are otherwise named buckets.
    name = (group or "").strip()
    if not name:
        return None
    folder = db.scalar(select(Folder).where(func.lower(Folder.name) == name.lower()))
    if not folder:
        folder = Folder(name=name, public_id=str(uuid.uuid4()))
        db.add(folder)
        db.flush()
    return folder.id


def _cluster_group_name(db: Session, cluster: KubeCluster) -> str | None:
    if cluster.folder_id is None:
        return None
    folder = db.get(Folder, cluster.folder_id)
    return folder.name if folder else None


def _cluster_read(db: Session, cluster: KubeCluster) -> KubeClusterRead:
    return KubeClusterRead(
        id=cluster.public_id,
        name=cluster.name,
        api_server_url=cluster.api_server_url,
        auth_method=cluster.auth_method,
        verify_tls=bool(cluster.verify_tls),
        default_namespace=cluster.default_namespace,
        group=_cluster_group_name(db, cluster),
        has_credentials=bool(cluster.encrypted_kubeconfig or cluster.encrypted_token),
        log_shipping_enabled=bool(getattr(cluster, "log_shipping_enabled", False)),
        log_namespaces=_cluster_namespaces(cluster),
        created_at=cluster.created_at,
    )


def _cluster_namespaces(cluster: KubeCluster) -> list[str]:
    try:
        value = json.loads(getattr(cluster, "log_namespaces_json", "") or "[]")
        return [str(n) for n in value] if isinstance(value, list) else []
    except (ValueError, TypeError):
        return []


def _cluster_conn(cluster: KubeCluster) -> dict:
    # Decrypt the stored secrets just-in-time for a single service call.
    return {
        "auth_method": cluster.auth_method,
        "api_server_url": cluster.api_server_url,
        "kubeconfig": decrypt_secret(cluster.encrypted_kubeconfig),
        "token": decrypt_secret(cluster.encrypted_token),
        "ca_cert": cluster.ca_cert,
        "verify_tls": bool(cluster.verify_tls),
    }


def _conn_from_create(payload: KubeClusterCreate) -> dict:
    return {
        "auth_method": _kube_auth_or_400(payload.auth_method),
        "api_server_url": payload.api_server_url,
        "kubeconfig": payload.kubeconfig,
        "token": payload.token,
        "ca_cert": payload.ca_cert,
        "verify_tls": payload.verify_tls,
    }


@router.get("/kube/clusters", response_model=list[KubeClusterRead])
def list_kube_clusters(_: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[KubeClusterRead]:
    # Cheap: reads rows only, never probes a cluster.
    clusters = db.scalars(select(KubeCluster).order_by(func.lower(KubeCluster.name))).all()
    return [_cluster_read(db, cluster) for cluster in clusters]


@router.put("/kube/clusters/{cluster_id}/log-shipping", response_model=KubeClusterRead)
def update_kube_log_shipping(cluster_id: str, payload: KubeLogShippingUpdate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> KubeClusterRead:
    # Enable/disable tailing this cluster's pod logs into Loki (namespaces=[] means all). The
    # background driver (app.main._k8s_log_shipping_loop) picks up enabled clusters each sweep.
    cluster = _cluster_or_404(db, cluster_id)
    namespaces = [str(n).strip() for n in (payload.namespaces or []) if str(n).strip()]
    cluster.log_shipping_enabled = bool(payload.enabled)
    cluster.log_namespaces_json = json.dumps(namespaces)
    db.commit()
    db.refresh(cluster)
    return _cluster_read(db, cluster)


@router.post("/kube/clusters", response_model=KubeClusterRead, status_code=status.HTTP_201_CREATED)
def create_kube_cluster(payload: KubeClusterCreate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> KubeClusterRead:
    method = _kube_auth_or_400(payload.auth_method)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cluster name is required")
    cluster = KubeCluster(
        public_id=str(uuid.uuid4()),
        name=name,
        auth_method=method,
        api_server_url=payload.api_server_url.strip(),
        encrypted_kubeconfig=encrypt_secret(payload.kubeconfig) if method == "kubeconfig" else "",
        encrypted_token=encrypt_secret(payload.token) if method == "token" else "",
        ca_cert=payload.ca_cert if method == "token" else "",
        verify_tls=payload.verify_tls,
        default_namespace=payload.default_namespace.strip(),
        folder_id=_resolve_group_to_folder_id(db, payload.group),
    )
    db.add(cluster)
    db.commit()
    db.refresh(cluster)
    return _cluster_read(db, cluster)


@router.post("/kube/clusters/test", response_model=KubeTestResult)
def test_kube_cluster(payload: KubeClusterCreate, _: dict = Depends(require_admin)) -> KubeTestResult:
    # Test the credentials as typed, without persisting anything. An unreachable cluster or bad
    # credentials is an expected outcome, reported as ok=false with HTTP 200 (like the DB console).
    conn = _conn_from_create(payload)
    try:
        version = kube.test_connection(conn)
        where = payload.api_server_url.strip() or "cluster"
        return KubeTestResult(ok=True, message=f"Connected to {where}", version=version)
    except KubeError as exc:
        return KubeTestResult(ok=False, message=str(exc))


@router.get("/kube/clusters/{cluster_id}", response_model=KubeClusterRead)
def get_kube_cluster(cluster_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> KubeClusterRead:
    return _cluster_read(db, _cluster_or_404(db, cluster_id))


@router.patch("/kube/clusters/{cluster_id}", response_model=KubeClusterRead)
def update_kube_cluster(cluster_id: str, payload: KubeClusterUpdate, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> KubeClusterRead:
    cluster = _cluster_or_404(db, cluster_id)
    data = payload.model_dump(exclude_unset=True)

    if "auth_method" in data and data["auth_method"] is not None:
        cluster.auth_method = _kube_auth_or_400(data["auth_method"])
    if "name" in data and data["name"] is not None:
        name = data["name"].strip()
        if not name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cluster name cannot be empty")
        cluster.name = name
    if "api_server_url" in data and data["api_server_url"] is not None:
        cluster.api_server_url = data["api_server_url"].strip()
    if "default_namespace" in data and data["default_namespace"] is not None:
        cluster.default_namespace = data["default_namespace"].strip()
    if "verify_tls" in data and data["verify_tls"] is not None:
        cluster.verify_tls = bool(data["verify_tls"])
    if "ca_cert" in data and data["ca_cert"] is not None:
        cluster.ca_cert = data["ca_cert"]
    # secret fields: only overwrite when a non-empty value is supplied, so an edit that leaves the
    # box blank keeps the stored credential instead of wiping it
    if data.get("kubeconfig"):
        cluster.encrypted_kubeconfig = encrypt_secret(data["kubeconfig"])
    if data.get("token"):
        cluster.encrypted_token = encrypt_secret(data["token"])
    if "group" in data:
        cluster.folder_id = _resolve_group_to_folder_id(db, data["group"])

    db.commit()
    db.refresh(cluster)
    return _cluster_read(db, cluster)


@router.delete("/kube/clusters/{cluster_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_kube_cluster(cluster_id: str, _: dict = Depends(require_admin), db: Session = Depends(get_db)) -> None:
    cluster = _cluster_or_404(db, cluster_id)
    db.delete(cluster)
    db.commit()


def _kube_call(func, *args):
    # Shared wrapper for the live reads/actions: a KubeError (unreachable cluster, API error, bad
    # credentials) is an expected outcome, surfaced as a 400 with a clean message, never a 500.
    try:
        return func(*args)
    except KubeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.get("/kube/clusters/{cluster_id}/overview", response_model=KubeOverview)
def kube_overview(cluster_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> KubeOverview:
    cluster = _cluster_or_404(db, cluster_id)
    return KubeOverview(**_kube_call(kube.overview, _cluster_conn(cluster)))


@router.get("/kube/clusters/{cluster_id}/nodes", response_model=list[KubeNode])
def kube_nodes(cluster_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[KubeNode]:
    cluster = _cluster_or_404(db, cluster_id)
    nodes = _kube_call(kube.list_nodes, _cluster_conn(cluster))
    # Link each node to a managed Server: primary match on internal IP, fallback on name==hostname
    # or name==alias (case-insensitive), so the UI can deep-link a node to its SSH host.
    servers = db.scalars(select(Server)).all()
    by_ip: dict[str, Server] = {}
    by_name: dict[str, Server] = {}
    for server in servers:
        if server.ip_address:
            by_ip.setdefault(server.ip_address.strip().lower(), server)
        if server.hostname:
            by_name.setdefault(server.hostname.strip().lower(), server)
        if server.alias:
            by_name.setdefault(server.alias.strip().lower(), server)
    for node in nodes:
        ip = (node.get("internal_ip") or "").strip().lower()
        match = by_ip.get(ip) if ip else None
        if match is None:
            match = by_name.get((node.get("name") or "").strip().lower())
        if match is not None:
            node["linked_server_id"] = match.public_id
            node["linked_server_alias"] = match.alias or match.hostname
    return [KubeNode(**node) for node in nodes]


@router.get("/kube/clusters/{cluster_id}/namespaces", response_model=list[str])
def kube_namespaces(cluster_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[str]:
    cluster = _cluster_or_404(db, cluster_id)
    return _kube_call(kube.list_namespaces, _cluster_conn(cluster))


@router.get("/kube/clusters/{cluster_id}/pods", response_model=list[KubePod])
def kube_pods(cluster_id: str, namespace: str = "", _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[KubePod]:
    cluster = _cluster_or_404(db, cluster_id)
    pods = _kube_call(kube.list_pods, _cluster_conn(cluster), namespace)
    return [KubePod(**pod) for pod in pods]


@router.get("/kube/clusters/{cluster_id}/pods/{namespace}/{pod}/logs", response_model=KubePodLogs)
def kube_pod_logs(cluster_id: str, namespace: str, pod: str, container: str = "", tail: int = 200, previous: bool = False, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> KubePodLogs:
    cluster = _cluster_or_404(db, cluster_id)
    result = _kube_call(kube.pod_logs, _cluster_conn(cluster), namespace, pod, container, tail, previous)
    return KubePodLogs(**result)


@router.get("/kube/clusters/{cluster_id}/deployments", response_model=list[KubeDeployment])
def kube_deployments(cluster_id: str, namespace: str = "", _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[KubeDeployment]:
    cluster = _cluster_or_404(db, cluster_id)
    deployments = _kube_call(kube.list_deployments, _cluster_conn(cluster), namespace)
    return [KubeDeployment(**dep) for dep in deployments]


@router.get("/kube/clusters/{cluster_id}/health", response_model=KubeHealth)
def kube_health(cluster_id: str, _: dict = Depends(require_user), db: Session = Depends(get_db)) -> KubeHealth:
    cluster = _cluster_or_404(db, cluster_id)
    return KubeHealth(**_kube_call(kube.health, _cluster_conn(cluster)))


@router.get("/kube/clusters/{cluster_id}/events", response_model=list[KubeEvent])
def kube_events(cluster_id: str, namespace: str = "", _: dict = Depends(require_user), db: Session = Depends(get_db)) -> list[KubeEvent]:
    cluster = _cluster_or_404(db, cluster_id)
    events = _kube_call(kube.list_events, _cluster_conn(cluster), namespace)
    return [KubeEvent(**ev) for ev in events]


@router.post("/kube/clusters/{cluster_id}/pods/{namespace}/{pod}/restart", response_model=ActionResult)
def kube_restart_pod(cluster_id: str, namespace: str, pod: str, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> ActionResult:
    cluster = _cluster_or_404(db, cluster_id)
    message = _kube_call(kube.restart_pod, _cluster_conn(cluster), namespace, pod)
    return ActionResult(ok=True, message=message)


@router.delete("/kube/clusters/{cluster_id}/pods/{namespace}/{pod}", response_model=ActionResult)
def kube_delete_pod(cluster_id: str, namespace: str, pod: str, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> ActionResult:
    cluster = _cluster_or_404(db, cluster_id)
    message = _kube_call(kube.delete_pod, _cluster_conn(cluster), namespace, pod)
    return ActionResult(ok=True, message=message)


@router.post("/kube/clusters/{cluster_id}/deployments/{namespace}/{name}/scale", response_model=ActionResult)
def kube_scale_deployment(cluster_id: str, namespace: str, name: str, payload: KubeScaleRequest, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> ActionResult:
    cluster = _cluster_or_404(db, cluster_id)
    message = _kube_call(kube.scale_deployment, _cluster_conn(cluster), namespace, name, payload.replicas)
    return ActionResult(ok=True, message=message)


@router.post("/kube/clusters/{cluster_id}/deployments/{namespace}/{name}/restart", response_model=ActionResult)
def kube_restart_deployment(cluster_id: str, namespace: str, name: str, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> ActionResult:
    cluster = _cluster_or_404(db, cluster_id)
    message = _kube_call(kube.restart_deployment, _cluster_conn(cluster), namespace, name)
    return ActionResult(ok=True, message=message)


@router.post("/kube/clusters/{cluster_id}/nodes/{name}/cordon", response_model=ActionResult)
def kube_cordon_node(cluster_id: str, name: str, payload: KubeCordonRequest, _: dict = Depends(require_admin_not_guest), db: Session = Depends(get_db)) -> ActionResult:
    cluster = _cluster_or_404(db, cluster_id)
    message = _kube_call(kube.cordon_node, _cluster_conn(cluster), name, payload.cordon)
    return ActionResult(ok=True, message=message)
